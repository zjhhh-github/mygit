
# -*- coding: utf-8 -*-

import sqlite3
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# 固定输入输出路径，保持和现有脚本一致，便于直接运行
输入Excel路径 = Path(r"C:\Users\LENOVO\Desktop\推广员信息.xlsx")
输出Excel路径 = Path(r"C:\Users\LENOVO\Desktop\推广员信息.xlsx")
数据库路径 = Path(r"C:\Users\LENOVO\Desktop\contact.db")
输出ID路径 = Path(r"C:\Users\LENOVO\Desktop\_输出结果_1.txt")
输出ID昵称路径 = Path(r"C:\Users\LENOVO\Desktop\_输出结果_1_含推广员昵称.txt")
输出ID昵称手机号路径 = Path(r"C:\Users\LENOVO\Desktop\_输出结果_1_含推广员昵称手机号.txt")
输出明细路径 = Path(r"C:\Users\LENOVO\Desktop\_输出结果_1_明细.csv")
未匹配标记 = "⚠️"
输入异常标记 = "❌"


def 自动定位推广员CSV():
    """
    自动定位推广员列表CSV，并自动识别可用编码。
    通过校验关键列名避免选错下载目录中的其它CSV。
    """
    关键列 = {"推广员id", "推广员昵称", "合伙宝妈-孩子中文全名"}
    候选编码 = ("utf-8-sig", "gb18030", "gbk", "utf-8", "utf-16", "utf-16le")
    下载目录 = Path(r"C:\Users\LENOVO\Downloads")
    候选文件 = sorted(下载目录.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)

    for 文件路径 in 候选文件:
        for 编码 in 候选编码:
            try:
                表头 = pd.read_csv(str(文件路径), encoding=编码, nrows=0)
            except Exception:
                # 该编码不可用或该文件非目标结构，尝试下一种
                continue
            if 关键列.issubset(set(表头.columns)):
                return str(文件路径), 编码

    raise FileNotFoundError("未找到包含关键列[推广员id/推广员昵称/合伙宝妈-孩子中文全名]的CSV文件")


def 读取输入文本行(文件路径):
    """兼容常见编码读取文本行。"""
    for 编码 in ("utf-8", "gb18030", "gbk"):
        try:
            with open(文件路径, "r", encoding=编码) as f:
                return [line.strip() for line in f.readlines() if line.strip()]
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("unknown", b"", 0, 1, "输入文件编码无法识别")


def 解析编号与孩子名(原始行):
    """
    从输入行中提取“编号 + 孩子中文全名”。
    行格式容错：允许前导符号、缩进、统计数字等噪音。
    """
    当前行 = 原始行.strip()
    # 业务约定：如果整行是“❌”，则直接按异常标记处理
    if 当前行 == 输入异常标记:
        return 输入异常标记, 输入异常标记

    # 兼容两种输入格式：
    # 1) 旧格式：000024-孙一可-287-287
    # 2) 新格式：000024-孙一可
    # 并允许行首存在缩进或其它噪音字符（例如“¿¿¿”）
    匹配 = re.search(r"(\d{6})-([^-]+?)(?:-\d+-\d+)?$", 当前行)
    if not 匹配:
        return "", ""
    编号 = 匹配.group(1).strip()
    孩子中文全名 = 匹配.group(2).strip()
    return 编号, 孩子中文全名


def 查询目标昵称(cursor, 编号, 孩子中文全名):
    """先精确匹配 remark，再使用 LIKE 兜底，避免误匹配。"""
    if not 编号 or not 孩子中文全名:
        return "未匹配"

    remark = f"{编号}-{孩子中文全名}"
    sql_精确 = "SELECT nick_name FROM contact WHERE remark = ?;"
    # 注意参数必须是单元素元组，避免 sqlite 参数绑定错误
    结果 = cursor.execute(sql_精确, (remark,)).fetchall()
    if len(结果) == 1:
        return str(结果[0][0]).strip()

    sql模糊 = "SELECT nick_name FROM contact WHERE remark LIKE ?;"
    结果模糊 = cursor.execute(sql模糊, (f"%{remark}%",)).fetchall()
    if len(结果模糊) == 1:
        return str(结果模糊[0][0]).strip()

    return "未匹配"


def 构建映射字典(推广员df):
    """
    构建两个映射：
    1) 推广员昵称 -> 推广员id
    2) 孩子中文全名 -> 推广员id（合伙宝妈字段，支持拆分多姓名）
    """
    昵称映射ID = {}
    孩子名映射ID = {}
    ID映射推广员昵称 = {}
    ID映射推广员手机号 = {}
    ID映射是否绑定 = {}

    def 规范化手机号(手机号文本):
        """去掉CSV浮点尾巴“.0”，并清理空白字符。"""
        文本 = str(手机号文本).strip()
        if not 文本 or 文本 == "nan":
            return ""
        if 文本.endswith(".0") and 文本[:-2].isdigit():
            return 文本[:-2]
        return 文本

    for _, row in 推广员df.iterrows():
        推广员id = str(row.get("推广员id", "")).strip()
        推广员昵称 = str(row.get("推广员昵称", "")).strip()
        孩子字段 = str(row.get("合伙宝妈-孩子中文全名", "")).strip()
        账户绑定手机号 = 规范化手机号(row.get("账户绑定手机号", ""))
        申请手机号 = 规范化手机号(row.get("申请手机号", ""))
        邀请人昵称 = str(row.get("邀请人昵称", "")).strip()
        邀请人姓名 = str(row.get("邀请人姓名", "")).strip()
        邀请人手机号 = 规范化手机号(row.get("邀请人手机号", ""))

        if 推广员id and 推广员id != "nan":
            if 推广员昵称 and 推广员昵称 != "nan":
                昵称映射ID[推广员昵称] = 推广员id
                # 反向保存“推广员id -> 推广员昵称”，用于结果输出昵称
                ID映射推广员昵称[推广员id] = 推广员昵称

            # 手机号优先取“申请手机号”，为空时回退到“账户绑定手机号”
            if 申请手机号:
                ID映射推广员手机号[推广员id] = 申请手机号
            elif 账户绑定手机号:
                ID映射推广员手机号[推广员id] = 账户绑定手机号
            # 只要有任一邀请人字段，即判定为“已绑定”
            if 邀请人昵称 and 邀请人昵称 != "nan":
                ID映射是否绑定[推广员id] = "✅"
            elif 邀请人姓名 and 邀请人姓名 != "nan":
                ID映射是否绑定[推广员id] = "✅"
            elif 邀请人手机号:
                ID映射是否绑定[推广员id] = "✅"
            else:
                ID映射是否绑定[推广员id] = 未匹配标记

            # 整字段先入库，确保完全一致时可直接命中
            if 孩子字段 and 孩子字段 != "nan":
                孩子名映射ID[孩子字段] = 推广员id
                # 常见分隔符拆分，提升“多孩子姓名”场景匹配率
                for 名称 in re.split(r"[、/，,\s]+", 孩子字段):
                    名称 = 名称.strip()
                    if 名称:
                        孩子名映射ID[名称] = 推广员id

    return 昵称映射ID, 孩子名映射ID, ID映射推广员昵称, ID映射推广员手机号, ID映射是否绑定


def 主流程():
    推广员CSV路径, CSV编码 = 自动定位推广员CSV()
    推广员df = pd.read_csv(推广员CSV路径, encoding=CSV编码)
    昵称映射ID, 孩子名映射ID, ID映射推广员昵称, ID映射推广员手机号, ID映射是否绑定 = 构建映射字典(推广员df)

    def 从孩子名映射匹配ID(孩子中文全名):
        """
        最小改动兜底：
        1) 先按整串精确匹配；
        2) 未命中时按常见分隔符拆分后逐个匹配。
        """
        if 孩子中文全名 in 孩子名映射ID:
            return 孩子名映射ID[孩子中文全名]
        for 名称 in re.split(r"[、/，,\s]+", 孩子中文全名):
            名称 = 名称.strip()
            if 名称 and 名称 in 孩子名映射ID:
                return 孩子名映射ID[名称]
        return ""

    def 从推广员昵称匹配ID(目标文本):
        """
        新增兜底：
        当“合伙宝妈-孩子中文全名”匹配失败时，使用“推广员昵称”再匹配一次。
        规则（保守）：
        1) 先做昵称精确匹配；
        2) 再做包含匹配，且仅在唯一命中时返回，避免误匹配。
        """
        目标文本 = str(目标文本 or "").strip()
        if not 目标文本:
            return ""

        if 目标文本 in 昵称映射ID:
            return 昵称映射ID[目标文本]

        候选ID = set()
        for 昵称, 推广员id in 昵称映射ID.items():
            昵称 = str(昵称).strip()
            if not 昵称:
                continue
            if 目标文本 in 昵称 or 昵称 in 目标文本:
                候选ID.add(推广员id)
        if len(候选ID) == 1:
            return next(iter(候选ID))
        return ""

    # 读取工作簿并仅处理活动工作表
    工作簿 = load_workbook(str(输入Excel路径))
    工作表 = 工作簿.active
    conn = sqlite3.connect(str(数据库路径))
    cursor = conn.cursor()
    try:
        # 第一行固定写表头
        工作表["A1"] = "合伙宝妈"
        工作表["B1"] = "合伙宝妈ID"
        工作表["C1"] = "上级合伙宝妈"
        工作表["D1"] = "上级合伙宝妈ID"
        工作表["E1"] = "手机号"
        工作表["F1"] = "是否绑定"

        for 行号 in range(2, 工作表.max_row + 1):
            A列文本 = str(工作表[f"A{行号}"].value or "").strip()
            C列文本 = str(工作表[f"C{行号}"].value or "").strip()

            # A列：仅填充推广员id到B列
            A编号, A孩子中文全名 = 解析编号与孩子名(A列文本)
            A推广员id = ""
            A是否绑定 = ""
            if A列文本:
                if A编号 == 输入异常标记 and A孩子中文全名 == 输入异常标记:
                    A推广员id = 输入异常标记
                    A是否绑定 = 输入异常标记
                elif A编号 and A孩子中文全名:
                    A昵称 = 查询目标昵称(cursor, A编号, A孩子中文全名)
                    A推广员id = 从孩子名映射匹配ID(A孩子中文全名)
                    # 按需求新增：孩子名未命中时，用“推广员昵称”做兜底匹配
                    if not A推广员id:
                        A推广员id = 从推广员昵称匹配ID(A孩子中文全名)
                    if not A推广员id and A昵称 in 昵称映射ID:
                        A推广员id = 昵称映射ID[A昵称]
                    if not A推广员id:
                        A推广员id = 未匹配标记
                    A是否绑定 = ID映射是否绑定.get(A推广员id, 未匹配标记)
                else:
                    # 非空但无法解析，按未匹配处理，便于人工排查
                    A推广员id = 未匹配标记
                    A是否绑定 = 未匹配标记
                工作表[f"B{行号}"] = A推广员id
                工作表[f"F{行号}"] = A是否绑定

            # C列：填充推广员id到D列、手机号到E列
            C编号, C孩子中文全名 = 解析编号与孩子名(C列文本)
            C推广员id = ""
            C推广员手机号 = ""
            if C列文本:
                if C编号 == 输入异常标记 and C孩子中文全名 == 输入异常标记:
                    C推广员id = 输入异常标记
                    C推广员手机号 = 输入异常标记
                elif C编号 and C孩子中文全名:
                    C昵称 = 查询目标昵称(cursor, C编号, C孩子中文全名)
                    C推广员id = 从孩子名映射匹配ID(C孩子中文全名)
                    # 按需求新增：孩子名未命中时，用“推广员昵称”做兜底匹配
                    if not C推广员id:
                        C推广员id = 从推广员昵称匹配ID(C孩子中文全名)
                    if not C推广员id and C昵称 in 昵称映射ID:
                        C推广员id = 昵称映射ID[C昵称]
                    if not C推广员id:
                        C推广员id = 未匹配标记
                    C推广员手机号 = ID映射推广员手机号.get(C推广员id, C推广员id)
                else:
                    # 非空但无法解析，按未匹配处理，便于人工排查
                    C推广员id = 未匹配标记
                    C推广员手机号 = 未匹配标记
                工作表[f"D{行号}"] = C推广员id
                工作表[f"E{行号}"] = C推广员手机号

            # 仅做控制台进度输出，避免中文符号导致编码异常
            if A列文本 or C列文本:
                print(f"第{行号}行已处理")
    finally:
        cursor.close()
        conn.close()
    # 输出为新文件，避免覆盖原始Excel；若文件被占用则自动回退文件名
    try:
        工作簿.save(str(输出Excel路径))
        print(f"已输出：{输出Excel路径}")
    except PermissionError:
        回退输出路径 = 输出Excel路径.with_name(f"{输出Excel路径.stem}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}{输出Excel路径.suffix}")
        工作簿.save(str(回退输出路径))
        print(f"输出文件被占用，已回退输出：{回退输出路径}")


if __name__ == "__main__":
    主流程()




