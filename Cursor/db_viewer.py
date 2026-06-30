"""
售前通讯录查询工具 - 数据库查看器

功能说明:
    - 从多个数据源加载联系人数据（AutoAgreeAddFriendTask、ContactAuthMsgTable、售前通讯录.txt）
    - 支持数据去重、搜索过滤、导出Excel
    - 自动备份数据库文件
    - 提供友好的图形界面
    - 支持报名情况查询和对外查询导出

性能优化:
    - 延迟导入 openpyxl，提升启动速度
    - 使用列索引常量，提升代码可维护性
    - 预加载数据映射表，减少数据库查询次数
    - 异步备份数据库，不阻塞主流程

数据结构:
    - 12列数据（不含序号）：
      对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、
      对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、
      来源总微信号、来源内部备注

作者: Cursor AI Assistant
版本: 2.1
最后更新: 2026-01-26
"""

# 启用DPI感知,解决Windows高DPI下界面模糊问题
import ctypes
try:
    # Windows 8.1及以上
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except:
    try:
        # Windows Vista/7
        ctypes.windll.user32.SetProcessDPIAware()
    except:
        pass

import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from typing import List, Tuple, Optional
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET
import threading
import os
import shutil
import time
import sys
import gc  # 垃圾回收，用于内存优化
# openpyxl 延迟导入（仅在导出时导入，提升启动速度）
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill


# ==================== 资源路径获取函数 ====================

def get_resource_path(relative_path: str) -> str:
    """
    获取资源文件的绝对路径（支持打包后的exe）
    
    在开发环境中，返回相对于当前文件的路径
    在打包后的exe中，返回临时解压目录中的资源路径
    
    Args:
        relative_path: 相对路径（如 "售前通讯录.txt"）
    
    Returns:
        资源文件的绝对路径
    """
    try:
        # PyInstaller 打包后的临时文件夹路径
        base_path = sys._MEIPASS
    except AttributeError:
        # 开发环境，使用当前文件所在目录
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)


# ==================== 常量定义 ====================

# 超微数据库路径配置（主路径 -> 备份路径）
PRIMARY_DB = r"C:\Users\LENOVO\AppData\Local\WxRobot\db_wxid_3iidhz1xmnta22.db3"
BACKUP_DB = r"C:\Users\LENOVO\Documents\wxid_3iidhz1xmnta22备份\db.db3"

# 内部数据库路径配置（网络路径 -> 本地备份）
CONTACT_NETWORK_BASE = r"X:\【技术】-专属共享文件夹\chatlog\内部专用号\chatlog"
CONTACT_LOCAL_BACKUP = r"C:\Users\LENOVO\Desktop\contact.db"

# 超微数据库网络备份路径
NETWORK_BACKUP_BASE = r"X:\【技术】-专属共享文件夹\售前超微备份"

# 导出密码配置（硬编码，仅用于数据导出验证）
EXPORT_PASSWORD = 'Cursor666#'
MAX_PASSWORD_ATTEMPTS = 3

# 时间转换常量（.NET Ticks 转 Unix 时间戳）
TICKS_TO_UNIX_EPOCH = 621355968000000000  # 0001-01-01 到 1970-01-01 的 Ticks 数
TICKS_PER_SECOND = 10000000  # 1秒 = 10,000,000 ticks (100纳秒)

# 数据列索引常量（values数组的索引，不含序号列）
# 使用常量避免硬编码索引，提升代码可维护性
COL_OBJ_NICK = 0              # 对象(昵称)
COL_OBJ_WXID = 1              # 对象(微信ID)
COL_OBJ_NUMBER = 2            # 对象(微信号)
COL_OBJ_TOTAL_NUMBER = 3      # 对象(总微信号)
COL_OBJ_TIME = 4              # 对象(添加时间)
COL_OBJ_INTERNAL_NOTE = 5     # 对象(内部备注)
COL_OBJ_IS_DELETE = 6         # 对象(是否删除)
COL_SOURCE_NICK = 7           # 来源(昵称)
COL_SOURCE_WXID = 8           # 来源(微信ID)
COL_SOURCE_NUMBER = 9         # 来源(微信号)
COL_SOURCE_TOTAL_NUMBER = 10  # 来源(总微信号)
COL_SOURCE_INTERNAL_NOTE = 11 # 来源(内部备注)


# ==================== 辅助函数 ====================

def clean_confirmation_mark(text: str) -> str:
    """
    清理【需二次确认】标记
    
    性能优化：使用单次 replace 替代多次操作
    
    Args:
        text: 待清理的文本
    
    Returns:
        清理后的文本
    """
    if not text:
        return ""
    return text.replace(" 【需二次确认】", "").replace("【需二次确认】", "").strip()


def clean_newlines(text: str) -> str:
    """
    去除文本中的换行符，用空格替换
    
    功能：
        - 将 \r\n、\n、\r 替换为空格
        - 保留多余空格（不做trim处理）
        - 用于清理昵称字段中的换行符
    
    Args:
        text: 待清理的文本
    
    Returns:
        清理后的文本（换行符已替换为空格）
    
    示例：
        "张三\n李四" -> "张三 李四"
        "王五\r\n赵六" -> "王五 赵六"
    """
    if not text:
        return ""
    # 按顺序替换：先处理 \r\n（避免被拆分），再处理单独的 \n 和 \r
    return text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')


# ==================== 数据库连接模块 ====================

def backup_to_network(source_path: str) -> bool:
    """
    备份数据库到网络路径（带时间标签）
    
    功能:
        - 检查源文件是否存在且可访问
        - 检查网络备份路径是否可访问
        - 生成带时间标签的备份文件名
        - 执行备份操作
    
    Args:
        source_path: 源数据库文件路径
    
    Returns:
        备份是否成功
        
    文件命名规则:
        原文件: db_wxid_3iidhz1xmnta22.db3
        备份文件: db_wxid_3iidhz1xmnta22_20260125_143025.db3
        时间格式: YYYYMMDD_HHMMSS
        
    注意:
        - 网络路径不可访问时静默失败
        - 备份失败不影响主流程
        - 使用异步线程执行，不阻塞数据加载
    """
    try:
        # 检查源文件是否存在
        if not source_path or not os.path.exists(source_path):
            return False
        
        # 检查源文件是否为文件（而非目录）
        if not os.path.isfile(source_path):
            return False
        
        # 检查网络备份路径是否可访问
        if not os.path.exists(NETWORK_BACKUP_BASE):
            return False
        
        # 检查网络路径是否为目录
        if not os.path.isdir(NETWORK_BACKUP_BASE):
            return False
        
        # 生成带时间标签的备份文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.basename(source_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_{timestamp}{ext}"
        
        # 构建完整的目标路径
        target_path = os.path.join(NETWORK_BACKUP_BASE, backup_filename)
        
        # 执行备份（复用现有的 backup_database 函数）
        success = backup_database(source_path, target_path)
        
        if success:
            print(f"[备份日志] 网络备份成功: {backup_filename}")
        else:
            print(f"[备份日志] 网络备份失败")
        
        return success
    
    except PermissionError as e:
        # 权限不足
        print(f"[备份日志] 网络备份权限不足: {e}")
        return False
    except OSError as e:
        # 文件系统错误（如网络路径不可访问）
        print(f"[备份日志] 网络备份文件系统错误: {e}")
        return False
    except Exception as e:
        # 其他未预期的错误
        print(f"[备份日志] 网络备份未知异常: {type(e).__name__}: {e}")
        return False


def backup_database(source_path: str, target_path: str) -> bool:
    """
    备份数据库文件（安全复制，保留元数据）
    
    Args:
        source_path: 源数据库文件路径
        target_path: 目标备份文件路径
    
    Returns:
        备份是否成功
        
    安全性:
        - 验证源文件存在且为文件
        - 自动创建目标目录
        - 使用 shutil.copy2 保留文件元数据
        - 静默处理异常，不影响主流程
        
    注意:
        - 如果目标文件已存在，会被覆盖
        - 备份失败不会抛出异常
    """
    try:
        # 验证源文件存在且不为空
        if not source_path or not os.path.exists(source_path):
            return False
        
        # 验证源文件是文件而非目录
        if not os.path.isfile(source_path):
            return False
        
        # 验证源文件大小（避免备份空文件或损坏文件）
        if os.path.getsize(source_path) == 0:
            return False
        
        # 创建目标文件夹（如果不存在）
        target_dir = os.path.dirname(target_path)
        if target_dir:
            os.makedirs(target_dir, exist_ok=True)
        
        # 复制文件（覆盖已存在的文件，保留元数据）
        shutil.copy2(source_path, target_path)
        
        # 验证备份文件是否成功创建
        if os.path.exists(target_path) and os.path.getsize(target_path) > 0:
            return True
        else:
            return False
    
    except PermissionError:
        # 权限不足，静默处理
        return False
    except shutil.Error:
        # 文件复制错误，静默处理
        return False
    except OSError:
        # 文件系统错误，静默处理
        return False
    except Exception:
        # 其他未预期的错误，静默处理
        return False


def get_contact_db_path() -> str:
    r"""
    动态获取contact.db数据库路径（智能日期排序）
    
    逻辑：
    1. 尝试访问网络路径 X:\【技术】-专属共享文件夹\chatlog\内部专用号\chatlog
    2. 筛选出尾部包含_日期格式的文件夹（如 xxx_20260115）
    3. 用_分割文件夹名称，提取尾部日期
    4. 按日期数值排序，取最新日期
    5. 构建数据库路径并验证
    
    Returns:
        contact.db 数据库完整路径（主路径或备用路径）
    """
    # 使用模块级常量
    primary_network_base = CONTACT_NETWORK_BASE
    backup_local_path = CONTACT_LOCAL_BACKUP
    
    # 尝试访问网络路径
    try:
        if os.path.exists(primary_network_base) and os.path.isdir(primary_network_base):
            # 获取所有文件夹
            all_folders = [f for f in os.listdir(primary_network_base) 
                          if os.path.isdir(os.path.join(primary_network_base, f))]
            
            # 筛选出尾部包含_日期的文件夹，并提取日期
            valid_folders = []
            for folder in all_folders:
                parts = folder.split('_')
                if len(parts) >= 2:
                    date_str = parts[-1]  # 取最后一节
                    # 验证日期格式（8位数字，YYYYMMDD）
                    if date_str.isdigit() and len(date_str) == 8:
                        try:
                            # 验证日期有效性
                            year = int(date_str[0:4])
                            month = int(date_str[4:6])
                            day = int(date_str[6:8])
                            if 1 <= month <= 12 and 1 <= day <= 31:
                                valid_folders.append((folder, int(date_str)))
                        except ValueError:
                            continue
            
            # 按日期排序，取最新的
            if valid_folders:
                valid_folders.sort(key=lambda x: x[1], reverse=True)
                latest_folder = valid_folders[0][0]
                primary_db_path = os.path.join(primary_network_base, latest_folder, 
                                              "db_storage", "contact", "contact.db")
                
                # 验证数据库文件是否存在
                if os.path.exists(primary_db_path) and os.path.isfile(primary_db_path):
                    return primary_db_path
    
    except (OSError, PermissionError, Exception):
        # 网络路径访问失败，静默处理
        pass
    
    # 网络路径不可访问或没有有效文件夹，返回备用路径
    return backup_local_path


def get_short_path(full_path: str, num_parts: int = 3) -> str:
    r"""
    截断路径，只显示最后N级
    
    Args:
        full_path: 完整路径
        num_parts: 保留的路径级数（默认3级）
    
    Returns:
        精简后的路径，格式: \part1\part2\part3
    
    示例:
        C:\Users\LENOVO\Desktop\contact.db -> \LENOVO\Desktop\contact.db
    """
    if not full_path:
        return ""
    
    # 统一使用反斜杠分割
    parts = full_path.replace('/', '\\').split('\\')
    
    # 过滤空字符串
    parts = [p for p in parts if p]
    
    # 取最后N级
    if len(parts) <= num_parts:
        return '\\' + '\\'.join(parts)
    else:
        return '\\' + '\\'.join(parts[-num_parts:])


class DatabaseReader:
    """数据库读取器,支持主备路径自动切换"""
    
    def __init__(self, primary_path: str, backup_path: str):
        """
        初始化数据库读取器
        
        Args:
            primary_path: 主数据库文件路径
            backup_path: 备用数据库文件路径
        """
        self.primary_path = Path(primary_path)
        self.backup_path = Path(backup_path)
        self.current_db_path: Optional[Path] = None
    
    def connect(self) -> Tuple[Optional[sqlite3.Connection], str]:
        """
        尝试连接数据库,优先使用主路径,失败时切换到备用路径
        
        Returns:
            (连接对象, 状态消息) 元组
        """
        # 尝试主路径
        conn, msg = self._try_connect(self.primary_path, is_primary=True)
        if conn:
            return conn, msg
        
        # 主路径失败,尝试备用路径
        conn, msg = self._try_connect(self.backup_path, is_primary=False)
        return conn, msg
    
    def _try_connect(self, db_path: Path, is_primary: bool) -> Tuple[Optional[sqlite3.Connection], str]:
        """
        尝试连接指定路径的数据库
        
        Args:
            db_path: 数据库文件路径
            is_primary: 是否为主路径
        
        Returns:
            (连接对象, 状态消息) 元组
        """
        path_type = "主路径" if is_primary else "备用路径"
        
        # 检查文件是否存在
        if not db_path.exists():
            return None, f"❌ {path_type}文件不存在: {db_path}"
        
        # 检查是否为文件(而非目录)
        if not db_path.is_file():
            return None, f"❌ {path_type}不是有效文件: {db_path}"
        
        # 尝试建立连接
        try:
            conn = sqlite3.connect(str(db_path))
            
            # 验证数据库可用性
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.close()
            
            self.current_db_path = db_path
            return conn, f"✓ 成功连接{path_type}: {db_path}"
            
        except sqlite3.DatabaseError as e:
            return None, f"❌ {path_type}数据库损坏或格式错误: {e}"
        
        except PermissionError:
            return None, f"❌ {path_type}权限不足,无法访问: {db_path}"
        
        except Exception as e:
            return None, f"❌ {path_type}连接失败: {type(e).__name__}: {e}"


def get_database_connection() -> Tuple[Optional[sqlite3.Connection], str]:
    """
    获取数据库连接的便捷函数
    
    Returns:
        (连接对象, 状态消息) 元组
    """
    # 创建读取器并连接（使用模块级常量）
    reader = DatabaseReader(PRIMARY_DB, BACKUP_DB)
    return reader.connect()


# ==================== 图形界面模块 ====================


class PasswordDialog(tk.Toplevel):
    """
    自定义密码输入对话框（现代化设计）
    
    功能:
        - 现代化、美观的界面设计
        - 更大的对话框和输入框，提升用户体验
        - 支持密码输入（显示为*）
        - 带图标的确定和取消按钮
        - 输入框聚焦边框高亮效果
        
    使用方法:
        dialog = PasswordDialog(parent, "标题", "提示信息")
        password = dialog.result  # 获取输入的密码，取消则为None
    """
    
    def __init__(self, parent, title, prompt):
        """
        初始化密码对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            prompt: 提示信息
        """
        super().__init__(parent)
        self.title(title)
        self.result = None
        
        # 设置对话框大小和位置（优化尺寸：560x450，确保所有元素清晰可见，包括按钮）
        self.geometry("560x450")
        self.resizable(False, False)
        
        # 设置背景色
        self.configure(bg="#f5f6fa")
        
        # 设置为模态对话框
        self.transient(parent)
        self.grab_set()
        
        # 居中显示
        self._center_window()
        
        # 创建界面
        self._create_widgets(prompt)
        
        # 等待对话框关闭
        self.wait_window()
    
    def _center_window(self):
        """将对话框居中显示在父窗口上"""
        self.update_idletasks()
        
        # 获取父窗口位置和大小
        parent_x = self.master.winfo_x()
        parent_y = self.master.winfo_y()
        parent_width = self.master.winfo_width()
        parent_height = self.master.winfo_height()
        
        # 计算对话框位置（居中）
        dialog_width = 560
        dialog_height = 450
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        
        self.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
    
    def _create_widgets(self, prompt):
        """创建对话框界面组件（现代化设计）"""
        # 主容器（白色卡片式设计）
        main_frame = tk.Frame(self, bg="#ffffff", padx=30, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # 图标和标题区域
        header_frame = tk.Frame(main_frame, bg="#ffffff")
        header_frame.pack(pady=(0, 15))
        
        # 锁图标（使用 Unicode 字符）
        icon_label = tk.Label(
            header_frame,
            text="🔒",
            font=("Segoe UI Emoji", 20),
            bg="#ffffff"
        )
        icon_label.pack()
        
        # 将提示信息拆分为“主提示”和“剩余次数”（避免挤占输入框空间）
        prompt_lines = [line.strip() for line in prompt.splitlines() if line.strip()]
        main_prompt = prompt_lines[0] if prompt_lines else prompt
        sub_prompt = prompt_lines[1] if len(prompt_lines) > 1 else ""
        
        # 主提示信息标签（增大字号）
        prompt_label = tk.Label(
            main_frame,
            text=main_prompt,
            font=("微软雅黑", 12),
            bg="#ffffff",
            fg="#2c3e50",
            justify=tk.CENTER,
            wraplength=440
        )
        prompt_label.pack(pady=(0, 6))
        
        # 副提示信息标签（剩余次数）
        if sub_prompt:
            sub_label = tk.Label(
                main_frame,
                text=sub_prompt,
                font=("微软雅黑", 10),
                bg="#ffffff",
                fg="#7f8c8d",
                justify=tk.CENTER,
                wraplength=440
            )
            sub_label.pack(pady=(0, 18))
        else:
            # 没有副提示时，保持间距
            spacer = tk.Frame(main_frame, height=18, bg="#ffffff")
            spacer.pack()
        
        # 密码输入框容器（明确边框，避免“看不见”）
        
        # 密码输入框（直接使用Entry，带明显边框，确保可见）
        self.password_entry = tk.Entry(
            main_frame,
            show='*',
            font=("微软雅黑", 14),
            width=40,
            relief=tk.SOLID,  # 实线边框，确保可见
            bd=2,  # 边框宽度2px
            fg="#2c3e50",
            bg="#ffffff",
            insertbackground="#3498db",
            insertwidth=3,
            highlightthickness=2,
            highlightbackground="#bdc3c7",
            highlightcolor="#3498db"
        )
        self.password_entry.pack(pady=(0, 30), ipady=12)  # ipady=12 使高度约44px
        self.password_entry.focus_set()
        
        # 绑定快捷键
        self.password_entry.bind("<Return>", lambda e: self._on_ok())
        self.password_entry.bind("<Escape>", lambda e: self._on_cancel())
        
        # 按钮容器（增加底部间距，确保按钮完全显示）
        button_frame = tk.Frame(main_frame, bg="#ffffff")
        button_frame.pack(pady=(0, 20))
        
        # 确定按钮（固定高度44px）
        ok_button = tk.Button(
            button_frame,
            text="✓ 确定",
            command=self._on_ok,
            font=("微软雅黑", 11, "bold"),
            bg="#3498db",
            fg="#ffffff",
            activebackground="#2980b9",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            cursor="hand2",
            bd=0,
            width=10,
            height=2  # 固定高度（约44px）
        )
        ok_button.pack(side=tk.LEFT, padx=8)
        
        # 取消按钮（固定高度44px）
        cancel_button = tk.Button(
            button_frame,
            text="✕ 取消",
            command=self._on_cancel,
            font=("微软雅黑", 11, "bold"),
            bg="#95a5a6",
            fg="#ffffff",
            activebackground="#7f8c8d",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            cursor="hand2",
            bd=0,
            width=10,
            height=2  # 固定高度（约44px）
        )
        cancel_button.pack(side=tk.LEFT, padx=8)
    
    def _on_ok(self):
        """确定按钮点击事件"""
        self.result = self.password_entry.get()
        self.destroy()
    
    def _on_cancel(self):
        """取消按钮点击事件"""
        self.result = None
        self.destroy()


class DatabaseViewer:
    """数据库查看器GUI"""
    
    def __init__(self, root: tk.Tk):
        """
        初始化数据库查看器
        
        Args:
            root: tkinter主窗口
        """
        self.root = root
        self.root.title("售前通讯录查询")
        
        # 设置主窗口背景色（浅灰色，提升视觉层次）
        self.root.configure(bg="#f5f6fa")
        
        # 设置最小尺寸
        self.root.minsize(1440, 900)
        
        # 设置初始全屏(最大化窗口)
        self.root.state('zoomed')
        
        # 数据库连接（原有业务逻辑）
        self.conn: Optional[sqlite3.Connection] = None
        self.current_db_path: str = ""
        
        # contact.db 数据库连接（新增，用于后续数据交叉）
        self.contact_conn: Optional[sqlite3.Connection] = None
        self.contact_db_path: str = ""
        
        # 性能优化：数据缓存（避免重复加载）
        self._data_cache_timestamp: Optional[float] = None  # 数据缓存时间戳
        self._cache_validity_seconds: int = 300  # 缓存有效期：5分钟
        
        # 首次启动标记（用于判断是否显示主数据库不存在的提醒）
        self.is_first_load: bool = True
        self.need_show_db_warning: bool = False  # 是否需要显示数据库警告
        
        # Toast提示相关
        self.toast_var = tk.StringVar(value="")
        self.toast_after: Optional[int] = None
        
        # Loading动画相关
        self.loading_frame: Optional[tk.Frame] = None
        self.loading_canvas: Optional[tk.Canvas] = None
        self.loading_bar_id: Optional[int] = None
        self.loading_animation_id: Optional[int] = None
        self.loading_position: int = 0
        self.loading_direction: int = 1  # 1表示向右，-1表示向左
        
        # 搜索功能相关
        self.all_data: List[dict] = []  # 存储完整数据集（用于搜索过滤）
        self.search_var = tk.StringVar(value="")  # 搜索框内容
        self.search_entry: Optional[tk.Entry] = None  # 搜索输入框引用
        self.search_count_label: Optional[tk.Label] = None  # 搜索结果计数标签
        self.empty_state_label: Optional[tk.Label] = None  # 空状态提示标签
        
        # 定时自动刷新相关（每5分钟刷新一次）
        self.auto_refresh_timer_id: Optional[int] = None  # 定时器ID
        self.auto_refresh_interval: int = 300000  # 5分钟 = 300000毫秒
        
        # 配置样式
        self._configure_styles()
        
        # 创建界面
        self._create_widgets()
        
        # 创建Loading组件(初始隐藏)
        self._create_loading_widget()
        
        # 首次加载数据
        self.load_data()
    
    def _center_window(self, width: int, height: int):
        """
        将窗口居中显示在屏幕中央
        
        Args:
            width: 窗口宽度
            height: 窗口高度
        """
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def _configure_styles(self):
        """配置界面样式主题"""
        style = ttk.Style()
        style.theme_use("clam")  # 使用更现代的主题
        
        # 配置Treeview样式 - 优化清晰度和列边界
        style.configure(
            "Custom.Treeview",
            background="#ffffff",
            foreground="#2c3e50",  # 深色文字,提高对比度
            rowheight=44,  # 固定行高44px（跨分辨率统一，不随DPI缩放）
            fieldbackground="#ffffff",
            font=("微软雅黑", 11),  # 增大字体
            borderwidth=1,  # 添加边框，让列边界更清晰
            relief="solid"  # 实线边框
        )
        
        # 配置Treeview表头样式 - 蓝色背景（优化清晰度和列边界）
        style.configure(
            "Custom.Treeview.Heading",
            background="#3498db",  # 标准蓝色背景
            foreground="#ffffff",
            font=("微软雅黑", 13, "bold"),  # 增大表头字体到13pt，提升清晰度
            relief="raised",  # 凸起效果，让列边界更明显
            borderwidth=2,  # 增加边框宽度，突出列分隔
            padding=10  # 增加内边距，提升可读性
        )
        
        # 鼠标悬停效果
        style.map(
            "Custom.Treeview",
            background=[("selected", "#E3F2FD")],
            foreground=[("selected", "#2c3e50")]  # 选中时保持深色文字
        )
        
        # 表头悬停效果
        style.map(
            "Custom.Treeview.Heading",
            background=[("active", "#2980b9")]
        )
        
        # 配置按钮样式
        style.configure(
            "Custom.TButton",
            font=("微软雅黑", 11),  # 增大按钮字体
            padding=10,
            relief=tk.FLAT
        )
        
        style.map(
            "Custom.TButton",
            background=[("active", "#2980b9"), ("!active", "#3498db")],
            foreground=[("active", "#ffffff"), ("!active", "#ffffff")]
        )
        
        # 配置滚动条样式（优化宽度和颜色，提升用户体验）
        style.configure(
            "Vertical.TScrollbar",
            background="#bdbdbd",  # 中灰色，更明显
            troughcolor="#f5f6fa",  # 槽道浅灰色
            borderwidth=0,
            arrowsize=16,  # 增大箭头尺寸
            width=20  # 增加滚动条宽度到20px，便于点击
        )
        
        style.map(
            "Vertical.TScrollbar",
            background=[
                ("active", "#9e9e9e"),  # 悬停时深灰色
                ("pressed", "#757575"),  # 按下时更深
                ("!active", "#bdbdbd")  # 默认中灰色
            ]
        )
    
    @staticmethod
    def convert_timestamp(timestamp: int) -> str:
        """
        将Unix时间戳(毫秒级)转换为可读时间格式
        
        Args:
            timestamp: Unix时间戳(毫秒)
        
        Returns:
            格式化的时间字符串,格式: YYYY-MM-DD HH:MM:SS
        """
        try:
            # 毫秒级时间戳需要除以1000转换为秒级
            return datetime.fromtimestamp(int(timestamp) / 1000).strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError, OSError, ZeroDivisionError):
            return ""
    
    @staticmethod
    def convert_dotnet_ticks(ticks: int) -> str:
        """
        将.NET DateTime.Ticks转换为可读时间格式
        
        Args:
            ticks: .NET DateTime.Ticks (从0001-01-01 00:00:00开始的100纳秒间隔数)
        
        Returns:
            格式化的时间字符串,格式: YYYY-MM-DD HH:MM:SS
            
        注意:
            - .NET Ticks 从 0001-01-01 00:00:00 开始计数
            - Unix 时间戳从 1970-01-01 00:00:00 开始计数
            - 需要减去两者之间的差值进行转换
        """
        try:
            if not ticks or ticks == 0:
                return ""
            
            # 转换为Unix时间戳(秒)，使用模块级常量
            unix_timestamp = (int(ticks) - TICKS_TO_UNIX_EPOCH) / TICKS_PER_SECOND
            
            # 转换为datetime对象并格式化
            return datetime.fromtimestamp(unix_timestamp).strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError, OSError, OverflowError):
            # 时间转换失败，返回空字符串（静默处理）
            return ""
    
    @staticmethod
    def convert_time_format(time_str: str) -> str:
        """
        转换时间格式：2025年3月 -> 2025-03-01 00:00:00
        
        Args:
            time_str: 原始时间字符串
        
        Returns:
            格式化后的时间字符串，如果不匹配则返回原值
        """
        if not time_str:
            return ""
        
        # 如果已经是标准格式（包含"-"和":"），直接返回
        if "-" in time_str and ":" in time_str:
            return time_str
        
        # 匹配"YYYY年M月"或"YYYY年MM月"格式
        import re
        match = re.match(r'(\d{4})年(\d{1,2})月', time_str.strip())
        if match:
            year = match.group(1)
            month = match.group(2).zfill(2)  # 补齐为两位数
            return f"{year}-{month}-01 00:00:00"
        
        # 如果不匹配，保留原始值
        return time_str
    
    @staticmethod
    def validate_remark_format(remark: str) -> bool:
        r"""
        验证 remark 格式是否符合要求
        
        格式要求：¿¿¿ + 连续6个数字 + - + 其他内容
        示例：¿¿¿000001-张三
        
        Args:
            remark: 待验证的 remark 字符串
        
        Returns:
            格式是否符合要求
        """
        if not remark or len(remark) < 10:
            return False
        
        # 检查前3个字符是否为 ¿¿¿
        if not remark.startswith("¿¿¿"):
            return False
        
        # 检查第4-9个字符（索引3-8）是否为6个数字
        digits = remark[3:9]
        if not digits.isdigit() or len(digits) != 6:
            return False
        
        # 检查第10个字符（索引9）是否为 -
        if len(remark) < 10 or remark[9] != '-':
            return False
        
        return True
    
    @staticmethod
    def parse_xml_content(content: str) -> Tuple[str, str]:
        """
        解析XML格式的content字段,提取sharecardusername和sharecardnickname
        
        Args:
            content: XML格式的字符串
        
        Returns:
            (sharecardusername, sharecardnickname) 元组,解析失败返回空串
        """
        if not content:
            return "", ""
        
        try:
            root = ET.fromstring(content)
            # sharecardusername 和 sharecardnickname 是 <msg> 标签的属性,不是子元素
            username = root.get('sharecardusername', '')
            nickname = root.get('sharecardnickname', '')
            return username or "", nickname or ""
        except ET.ParseError:
            # XML解析失败
            return "", ""
        except Exception:
            # 其他异常
            return "", ""
    
    def _create_widgets(self):
        """创建界面组件"""
        
        # 顶部信息栏（白色卡片式设计）
        top_frame = tk.Frame(self.root, bg="#ffffff", padx=20, pady=15)
        top_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        # 左侧搜索区域
        search_frame = tk.Frame(top_frame, bg="#ffffff")
        search_frame.pack(side=tk.LEFT)
        
        # 搜索标签（去掉图标，使用中文冒号）
        search_label = tk.Label(
            search_frame,
            text="搜索：",
            font=("微软雅黑", 12, "bold"),
            bg="#ffffff",
            fg="#2c3e50"
        )
        search_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # 搜索输入框容器（用于添加边框效果）
        entry_container = tk.Frame(
            search_frame,
            bg="#bdc3c7",  # 边框颜色
            bd=0,
            highlightthickness=1,
            highlightbackground="#bdc3c7",
            highlightcolor="#3498db"  # 聚焦时边框颜色
        )
        entry_container.pack(side=tk.LEFT, padx=(0, 12))
        
        # 搜索输入框（现代化样式，无占位文本）
        self.search_entry = tk.Entry(
            entry_container,
            textvariable=self.search_var,
            font=("微软雅黑", 12),
            width=35,  # 增加宽度
            relief=tk.FLAT,
            bd=0,
            fg="#1a1a1a",  # 更醒目的黑色
            bg="#ffffff",
            insertbackground="#3498db"  # 光标颜色
        )
        self.search_entry.pack(padx=2, pady=2, ipady=6)  # 增加内边距，提升高度
        
        # 边框高亮效果（聚焦/失焦）
        def on_focus_in(event):
            # 聚焦时边框高亮
            entry_container.config(highlightbackground="#3498db", highlightthickness=2)
        
        def on_focus_out(event):
            # 失焦时恢复边框
            entry_container.config(highlightbackground="#bdc3c7", highlightthickness=1)
        
        self.search_entry.bind("<FocusIn>", on_focus_in)
        self.search_entry.bind("<FocusOut>", on_focus_out)
        
        # 绑定实时搜索（边输入边搜索）
        self.search_var.trace_add("write", lambda *args: self._on_search_change())
        
        # 清空按钮（现代化样式，添加图标）
        clear_btn = tk.Button(
            search_frame,
            text="✕ 清空",
            command=self._on_clear_button_click,
            font=("微软雅黑", 11, "bold"),
            bg="#95a5a6",
            fg="#ffffff",
            activebackground="#7f8c8d",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            padx=18,
            pady=10,  # 增加高度与搜索框对齐
            cursor="hand2",
            bd=0
        )
        clear_btn.pack(side=tk.LEFT, padx=(0, 12))
        
        # 搜索结果计数标签（在清空按钮右侧）
        self.search_count_label = tk.Label(
            search_frame,
            text="",
            font=("微软雅黑", 12, "bold"),
            fg="#2ecc71",  # 绿色，与底部状态一致
            bg="#ffffff"
        )
        self.search_count_label.pack(side=tk.LEFT)
        
        # 右侧按钮区域
        # 刷新按钮（右对齐，自定义样式）
        refresh_btn = tk.Button(
            top_frame,
            text="🔄 刷新数据",
            command=self.load_data,
            font=("微软雅黑", 11),
            bg="#3498db",
            fg="#ffffff",
            activebackground="#2980b9",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            bd=0
        )
        refresh_btn.pack(side=tk.RIGHT)
        
        # 到底按钮
        bottom_btn = tk.Button(
            top_frame,
            text="↓ 到底",
            command=self._scroll_to_bottom,
            font=("微软雅黑", 11),
            bg="#27ae60",
            fg="#ffffff",
            activebackground="#229954",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            bd=0
        )
        bottom_btn.pack(side=tk.RIGHT, padx=(0, 10))
        
        # 到顶按钮
        top_btn = tk.Button(
            top_frame,
            text="↑ 到顶",
            command=self._scroll_to_top,
            font=("微软雅黑", 11),
            bg="#27ae60",
            fg="#ffffff",
            activebackground="#229954",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            bd=0
        )
        top_btn.pack(side=tk.RIGHT, padx=(0, 10))
        
        # 导出按钮（在到顶按钮右侧）
        export_btn = tk.Button(
            top_frame,
            text="📥 导出数据",
            command=self._export_data,
            font=("微软雅黑", 11),
            bg="#e67e22",
            fg="#ffffff",
            activebackground="#d35400",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            bd=0
        )
        export_btn.pack(side=tk.RIGHT, padx=(0, 10))
        
        # 表格区域（白色背景容器）
        table_container = tk.Frame(self.root, bg="#ffffff", padx=5, pady=5)
        table_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        
        # 表格内部框架
        table_frame = tk.Frame(table_container, bg="#ffffff")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建Treeview表格（新顺序：序号、对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、来源总微信号、来源内部备注）
        columns = ("row_num", "nick", "wxid", "number", "total_number", "time", "obj_internal_note", "is_delete", "source_nickname", "source_username", "source_number", "source_total_number", "source_internal_note")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Custom.Treeview"
        )
        
        # 设置列标题
        self.tree.heading("row_num", text="序号")
        self.tree.heading("nick", text="对象(昵称)")
        self.tree.heading("wxid", text="对象(微信ID)")
        self.tree.heading("number", text="对象(微信号)")
        self.tree.heading("total_number", text="对象(总微信号)")
        self.tree.heading("time", text="对象(添加时间)")
        self.tree.heading("obj_internal_note", text="对象(内部备注)")
        self.tree.heading("is_delete", text="对象(是否删除)")
        self.tree.heading("source_nickname", text="来源(昵称)")
        self.tree.heading("source_username", text="来源(微信ID)")
        self.tree.heading("source_number", text="来源(微信号)")
        self.tree.heading("source_total_number", text="来源(总微信号)")
        self.tree.heading("source_internal_note", text="来源(内部备注)")
        
        # 设置列宽(适配1600宽度,用户可手动调整)
        self.tree.column("row_num", width=60, minwidth=120, anchor=tk.CENTER)
        self.tree.column("nick", width=140, minwidth=120, anchor=tk.W)
        self.tree.column("wxid", width=170, minwidth=150, anchor=tk.W)
        self.tree.column("number", width=130, minwidth=110, anchor=tk.W)
        self.tree.column("total_number", width=130, minwidth=110, anchor=tk.W)
        self.tree.column("time", width=140, minwidth=130, anchor=tk.W)
        self.tree.column("obj_internal_note", width=200, minwidth=150, anchor=tk.W)
        self.tree.column("is_delete", width=100, minwidth=80, anchor=tk.CENTER)
        self.tree.column("source_nickname", width=180, minwidth=140, anchor=tk.W)
        self.tree.column("source_username", width=180, minwidth=140, anchor=tk.W)
        self.tree.column("source_number", width=130, minwidth=110, anchor=tk.W)
        self.tree.column("source_total_number", width=130, minwidth=110, anchor=tk.W)
        self.tree.column("source_internal_note", width=200, minwidth=150, anchor=tk.W)
        
        # 添加交替行颜色
        self.tree.tag_configure("oddrow", background="#F8F9FA")
        self.tree.tag_configure("evenrow", background="white")
        
        # 添加警告行样式（用于标记需要二次确认的数据）
        self.tree.tag_configure("warning_odd", background="#F8F9FA", foreground="#e74c3c")
        self.tree.tag_configure("warning_even", background="white", foreground="#e74c3c")
        
        # 添加新数据标识样式（蓝色，用于调试）
        self.tree.tag_configure("new_data_odd", background="#F8F9FA", foreground="#3498db")
        self.tree.tag_configure("new_data_even", background="white", foreground="#3498db")
        
        # 添加已删除标记样式（红色，用于标记已删除的联系人）
        self.tree.tag_configure("deleted_odd", background="#F8F9FA", foreground="#e74c3c")
        self.tree.tag_configure("deleted_even", background="white", foreground="#e74c3c")
        
        # 添加文本文件数据标识样式（黑色，用于标记来自售前通讯录.txt的数据）
        self.tree.tag_configure("txt_file_odd", background="#F8F9FA", foreground="#2c3e50")
        self.tree.tag_configure("txt_file_even", background="white", foreground="#2c3e50")
        
        # 添加占位行样式（灰色、斜体，用于搜索无结果时的提示）
        self.tree.tag_configure("placeholder", background="#f5f6fa", foreground="#95a5a6", font=("微软雅黑", 12, "italic"))
        
        # 绑定单击事件
        self.tree.bind("<ButtonRelease-1>", self._on_cell_click)
        
        # 绑定键盘快捷键
        self.tree.bind("<Home>", lambda e: self._scroll_to_top())
        self.tree.bind("<End>", lambda e: self._scroll_to_bottom())
        
        # 添加竖向滚动条（移除横向滚动条，优化美观度）
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
        
        # 布局表格和滚动条（只保留竖向滚动条）
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 配置表格区域权重,使其可伸缩
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        
        # 空状态提示标签（居中显示在表格上方，初始隐藏）
        self.empty_state_label = tk.Label(
            table_frame,
            text="🔍 未找到匹配的数据，请尝试其他关键词",
            font=("微软雅黑", 16, "bold"),
            fg="#95a5a6",
            bg="#ffffff"
        )
        # 使用 place 布局，居中显示（初始隐藏）
        # 注意：place 会覆盖在 grid 布局的表格上方
        self.empty_state_label.place(in_=table_frame, relx=0.5, rely=0.5, anchor="center")
        self.empty_state_label.place_forget()
        
        # 底部状态栏（白色卡片式设计）
        status_frame = tk.Frame(self.root, bg="#ffffff", padx=20, pady=15)
        status_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        # 状态栏内部垂直布局容器
        status_container = tk.Frame(status_frame, bg="#ffffff")
        status_container.pack(side=tk.LEFT, fill=tk.BOTH)
        
        # 第一行：加载状态
        self.status_label = tk.Label(
            status_container,
            text="就绪",
            font=("微软雅黑", 12, "bold"),
            fg="#2ecc71",
            bg="#ffffff",
            anchor=tk.W
        )
        self.status_label.pack(side=tk.TOP, anchor=tk.W)
        
        # 第二行：原有数据库路径（初始隐藏）
        self.main_db_label = tk.Label(
            status_container,
            text="",
            font=("微软雅黑", 10),
            fg="#7f8c8d",  # 灰色，与绿色加载状态区分
            bg="#ffffff",
            anchor=tk.W
        )
        self.main_db_label.pack(side=tk.TOP, anchor=tk.W, pady=(5, 0))
        
        # 第三行：contact.db 路径（初始隐藏）
        self.contact_db_label = tk.Label(
            status_container,
            text="",
            font=("微软雅黑", 10),
            fg="#7f8c8d",  # 灰色，与绿色加载状态区分
            bg="#ffffff",
            anchor=tk.W
        )
        self.contact_db_label.pack(side=tk.TOP, anchor=tk.W, pady=(2, 0))
        
        # Toast提示标签（初始隐藏，参考query_tool.py）
        self.toast_label = tk.Label(
            self.root,
            textvariable=self.toast_var,
            bg="#2ecc71",
            fg="#ffffff",
            font=("微软雅黑", 12, "bold"),
            padx=25,
            pady=12,
            relief=tk.FLAT,
            bd=0
        )
    
    def _on_cell_click(self, event):
        """
        处理单元格点击事件，复制内容到剪贴板
        
        复制规则（按优先级）：
            1. ❌ 序号列 → 不允许复制
            2. ❌ 空单元格 → 不允许复制
            3. ❌ "已删除"单元格 → 不允许复制
            4. ❌ "待填写"、"搜索"单元格 → 不允许复制（占位符文本）
            5. ✅ 【需二次确认】单元格 → 只复制原始值（去掉标记）
        
        Args:
            event: 点击事件对象
        """
        # 获取点击的行和列
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        column = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        
        if not row_id:
            return
        
        # 规则1：禁止复制序号列（第1列，column="#1"）
        if column == "#1":
            return  # 序号列不执行复制操作
        
        # 检查是否为占位行（搜索无结果提示行）
        row_tags = self.tree.item(row_id, "tags")
        if "placeholder" in row_tags:
            return  # 占位行不执行复制操作
        
        # 获取单元格内容
        column_index = int(column.replace("#", "")) - 1
        row_values = self.tree.item(row_id, "values")
        
        if column_index < len(row_values):
            cell_content = str(row_values[column_index])
            
            # 规则2：判断单元格是否为空（包括空字符串、空格、None、N/A、null等）
            if not cell_content or cell_content.strip() == "" or cell_content.strip().lower() in ["none", "null", "n/a", "-"]:
                return  # 空值不执行复制操作
            
            # 规则2：判断是否为纯标记的空值（只包含【需二次确认】，无实际内容）
            if cell_content.strip() == "【需二次确认】":
                return  # 纯标记的空值不执行复制操作
            
            # 规则3：判断是否为"已删除"
            if cell_content.strip() == "已删除":
                return  # "已删除"不执行复制操作
            
            # 规则4：判断是否为特定占位符文本（"待填写"、"搜索"）
            if cell_content.strip() in ["待填写", "搜索"]:
                return  # 占位符文本不执行复制操作
            
            # 规则5：清理【需二次确认】标记文字，只复制实际内容
            clean_content = cell_content.replace(" 【需二次确认】", "").replace("【需二次确认】", "").strip()
            
            # 二次检查：如果清理后没有内容，不执行复制
            if not clean_content:
                return
            
            # 复制到剪贴板（复制清理后的内容）
            self.root.clipboard_clear()
            self.root.clipboard_append(clean_content)
            
            # 显示Toast提示（显示清理后的内容）
            self._show_toast(clean_content)
    
    def _scroll_to_top(self):
        """滚动到表格顶部"""
        self.tree.yview_moveto(0)
    
    def _scroll_to_bottom(self):
        """滚动到表格底部"""
        self.tree.yview_moveto(1)
    
    def _export_data(self):
        """
        导出数据到Excel文件（需要密码验证）
        
        功能:
            - 密码验证（最多3次尝试）
            - 使用自定义对话框，确保标题和内容完整显示
            - 验证通过后同时导出两个Excel表：完整导出表 + 对外查询表
            
        安全性:
            - 密码硬编码在常量中
            - 限制尝试次数防止暴力破解
        """
        for attempt in range(MAX_PASSWORD_ATTEMPTS):
            # 弹出自定义密码输入对话框
            prompt = f"请输入导出密码\n\n剩余尝试次数：{MAX_PASSWORD_ATTEMPTS - attempt}"
            dialog = PasswordDialog(self.root, "密码验证 - 数据导出", prompt)
            password = dialog.result
            
            # 用户取消输入
            if password is None:
                return
            
            # 验证密码（使用模块级常量）
            if password == EXPORT_PASSWORD:
                # 密码正确，执行两个导出操作
                self._do_export_all()
                return
            else:
                # 密码错误，提示用户
                if attempt < MAX_PASSWORD_ATTEMPTS - 1:
                    messagebox.showerror(
                        "密码错误",
                        f"密码错误！请重试。\n\n剩余尝试次数：{MAX_PASSWORD_ATTEMPTS - attempt - 1}"
                    )
                else:
                    messagebox.showerror(
                        "密码错误",
                        "密码错误次数过多，导出操作已取消。"
                    )
    
    def _do_export_all(self):
        """
        执行所有导出操作（完整导出 + 对外查询导出）
        
        功能:
            - 同时导出两个Excel文件
            - 统一处理成功/失败提示
        """
        # 检查是否有数据
        if not self.all_data or len(self.all_data) == 0:
            messagebox.showwarning("导出失败", "没有可导出的数据！")
            return
        
        # 生成时间戳（两个文件使用相同的时间戳）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 执行两个导出操作
        result1, filename1, error1 = self._do_export_full(timestamp)
        result2, filename2, error2 = self._do_export_external_query(timestamp)
        
        # 统一处理结果
        if result1 and result2:
            # 两个都成功
            messagebox.showinfo(
                "导出成功",
                f"数据已成功导出！\n\n"
                f"已导出以下文件：\n"
                f"1. {filename1}（完整数据，{len(self.all_data)} 条）\n"
                f"2. {filename2}（对外查询，{len(self.all_data)} 条）\n\n"
                f"保存位置：当前目录"
            )
        elif result1 and not result2:
            # 完整导出成功，对外查询导出失败
            messagebox.showwarning(
                "部分导出成功",
                f"完整导出成功：{filename1}\n\n"
                f"对外查询导出失败：{error2}"
            )
        elif not result1 and result2:
            # 完整导出失败，对外查询导出成功
            messagebox.showwarning(
                "部分导出成功",
                f"对外查询导出成功：{filename2}\n\n"
                f"完整导出失败：{error1}"
            )
        else:
            # 两个都失败
            messagebox.showerror(
                "导出失败",
                f"完整导出失败：{error1}\n\n"
                f"对外查询导出失败：{error2}"
            )
    
    def _do_export_full(self, timestamp: str) -> Tuple[bool, str, str]:
        """
        执行完整数据导出操作
        
        功能:
            - 将当前显示的数据导出为Excel文件
            - 排除序号列
            - 清理【需二次确认】标记
            - 转换删除状态（已删除→✅，未删除→❌）
            - 设置专业的Excel格式（字体、对齐、行高、列宽）
            - 冻结表头
        
        Args:
            timestamp: 时间戳字符串（格式：YYYYMMDD_HHMMSS）
            
        Returns:
            (是否成功, 文件名, 错误信息) 元组
        """
        # 延迟导入 openpyxl（仅在导出时导入，提升启动速度）
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        filename = f"售前通讯录导出_{timestamp}.xlsx"
        
        try:
            # 安全性检查：确保文件名不包含路径分隔符（防止路径遍历）
            if '/' in filename or '\\' in filename or '..' in filename:
                return False, filename, "文件名包含非法字符！"
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "售前通讯录"
            
            # 定义表头（排除序号列，新增"对象(内部备注)"）
            headers = [
                "对象(昵称)", "对象(微信ID)", "对象(微信号)", "对象(总微信号)",
                "对象(添加时间)", "对象(内部备注)", "对象(是否删除)", "来源(昵称)", "来源(微信ID)",
                "来源(微信号)", "来源(总微信号)", "来源(内部备注)"
            ]
            
            # 写入表头
            ws.append(headers)
            
            # 冻结表头（第一行）
            ws.freeze_panes = 'A2'
            
            # 设置表头样式（微软雅黑、11号、加粗、居中显示、黑色文字、灰色-25%背景）
            header_font = Font(name="微软雅黑", bold=True, color="000000", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # 灰色-25%，背景2
            
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
            
            # 写入数据（排除序号列）
            # 性能优化：使用列表推导式批量处理
            for item in self.all_data:
                values = item['values']
                # 清理【需二次确认】标记，并转换删除状态
                cleaned_values = [
                    # 特殊处理：转换删除状态
                    ("✅" if clean_confirmation_mark(str(val)) == "已删除" else "❌") 
                    if idx == COL_OBJ_IS_DELETE 
                    else clean_confirmation_mark(str(val))
                    for idx, val in enumerate(values)
                ]
                ws.append(cleaned_values)
            
            # 统一设置列宽为 19.5
            for col_idx in range(1, len(headers) + 1):
                col_letter = chr(64 + col_idx)  # A, B, C, ...
                ws.column_dimensions[col_letter].width = 19.5
            
            # 统一设置行高为 19.5（包括表头和数据行）
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 19.5
            
            # 性能优化：预创建样式对象，使用 iter_rows 批量设置
            data_font = Font(name="微软雅黑", color="000000", size=11)
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            # 批量设置数据行样式（使用 iter_rows 提升性能）
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
            
            # 保存文件到当前目录
            wb.save(filename)
            
            # 性能优化：释放工作簿对象，减少内存占用
            del wb
            del ws
            gc.collect()
            
            return True, filename, ""
            
        except PermissionError:
            return False, filename, f"文件 {filename} 正在被其他程序占用！请关闭该文件后重试。"
        except Exception as e:
            return False, filename, f"{type(e).__name__}: {e}"
    
    def _do_export_external_query(self, timestamp: str) -> Tuple[bool, str, str]:
        """
        执行对外查询数据导出操作
        
        功能:
            - 导出4列精简数据：对象(微信号)、来源(微信号)、添加时间、报名情况
            - 报名情况判断：直接使用已加载的"对象(内部备注)"列判断
            - 设置专业的Excel格式（字体、对齐、行高、列宽）
            - 冻结表头
        
        Args:
            timestamp: 时间戳字符串（格式：YYYYMMDD_HHMMSS）
            
        Returns:
            (是否成功, 文件名, 错误信息) 元组
        """
        # 延迟导入 openpyxl（仅在导出时导入，提升启动速度）
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        filename = f"售前通讯录对外查询导出_{timestamp}.xlsx"
        
        try:
            # 安全性检查：确保文件名不包含路径分隔符（防止路径遍历）
            if '/' in filename or '\\' in filename or '..' in filename:
                return False, filename, "文件名包含非法字符！"
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "对外查询"
            
            # 定义表头（4列）
            headers = [
                "对象(微信号)",
                "来源(微信号)",
                "添加时间",
                "报名情况"
            ]
            
            # 写入表头
            ws.append(headers)
            
            # 冻结表头（第一行）
            ws.freeze_panes = 'A2'
            
            # 设置表头样式（微软雅黑、11号、加粗、居中显示、黑色文字、灰色-25%背景）
            header_font = Font(name="微软雅黑", bold=True, color="000000", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # 灰色-25%，背景2
            
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
            
            # 统计报名情况
            registered_count = 0
            unregistered_count = 0
            
            print(f"\n[对外查询导出] 开始导出，直接使用已加载的\"对象(内部备注)\"列判断报名情况")
            
            # 写入数据
            for idx, item in enumerate(self.all_data):
                values = item['values']
                
                # 使用列索引常量和优化的清理函数提取数据
                obj_total_number = clean_confirmation_mark(str(values[COL_OBJ_TOTAL_NUMBER]))  # 对象(总微信号)
                source_total_number = clean_confirmation_mark(str(values[COL_SOURCE_TOTAL_NUMBER]))  # 来源(总微信号)
                add_time = str(values[COL_OBJ_TIME]).strip()  # 添加时间
                obj_internal_note = str(values[COL_OBJ_INTERNAL_NOTE]).strip()  # 对象(内部备注)
                
                # 判断报名情况：直接使用"对象(内部备注)"列的内容
                registration_status = "未报名"
                if obj_internal_note and self.validate_remark_format(obj_internal_note):
                    registration_status = "已报名"
                    registered_count += 1
                else:
                    unregistered_count += 1
                
                # 调试日志：输出前3条数据
                if idx < 3:
                    print(f"[调试] 第{idx+1}条 - 对象内部备注: {obj_internal_note}, 报名情况: {registration_status}")
                
                # 添加到表格
                ws.append([obj_total_number, source_total_number, add_time, registration_status])
            
            # 输出统计信息
            print(f"\n[对外查询导出] 报名情况统计:")
            print(f"  - 已报名: {registered_count} 条")
            print(f"  - 未报名: {unregistered_count} 条")
            print(f"  - 总计: {len(self.all_data)} 条\n")
            
            # 统一设置列宽为 19.5
            for col_idx in range(1, len(headers) + 1):
                col_letter = chr(64 + col_idx)  # A, B, C, ...
                ws.column_dimensions[col_letter].width = 19.5
            
            # 统一设置行高为 19.5（包括表头和数据行）
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 19.5
            
            # 性能优化：预创建样式对象，使用 iter_rows 批量设置
            data_font = Font(name="微软雅黑", color="000000", size=11)
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            # 批量设置数据行样式（使用 iter_rows 提升性能）
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
            
            # 保存文件到当前目录
            wb.save(filename)
            
            # 性能优化：释放工作簿对象，减少内存占用
            del wb
            del ws
            gc.collect()
            
            return True, filename, ""
            
        except PermissionError:
            return False, filename, f"文件 {filename} 正在被其他程序占用！请关闭该文件后重试。"
        except Exception as e:
            return False, filename, f"{type(e).__name__}: {e}"
    
    def _show_toast(self, content: str, duration_ms: int = 1500):
        """
        显示Toast提示消息（参考query_tool.py的实现）
        
        Args:
            content: 要显示的内容
            duration_ms: 显示时长（毫秒）
        """
        # 取消之前的定时任务
        if self.toast_after is not None:
            self.root.after_cancel(self.toast_after)
            self.toast_after = None
        
        # 内容限制长度
        display_content = content if len(content) <= 50 else content[:47] + "..."
        
        # 设置Toast文本（添加成功图标）
        self.toast_var.set(f"✓ 已复制：{display_content}")
        
        # 恢复默认样式（绿色背景）
        self.toast_label.config(bg="#2ecc71")
        
        # 使用place布局，居中显示，位置稍微靠下（参考query_tool.py）
        self.toast_label.place(relx=0.5, rely=0.15, anchor="n")
        
        # 设置自动隐藏
        self.toast_after = self.root.after(duration_ms, self._hide_toast)
    
    def _show_warning_toast(self, content: str, duration_ms: int = 5000):
        """
        显示警告Toast提示（红色背景，更醒目）
        
        Args:
            content: 要显示的内容
            duration_ms: 显示时长（毫秒）
        """
        # 取消之前的定时任务
        if self.toast_after is not None:
            self.root.after_cancel(self.toast_after)
            self.toast_after = None
        
        # 设置Toast文本（添加警告图标）
        self.toast_var.set(f"⚠ {content}")
        
        # 设置红色背景（更醒目）
        self.toast_label.config(bg="#e74c3c")
        
        # 使用place布局，居中显示，位置稍微靠下
        self.toast_label.place(relx=0.5, rely=0.15, anchor="n")
        
        # 设置自动隐藏
        self.toast_after = self.root.after(duration_ms, self._hide_toast)
    
    def _hide_toast(self):
        """隐藏Toast提示"""
        self.toast_label.place_forget()
        self.toast_after = None
    
    def _on_search_change(self):
        """
        搜索框内容变化时的处理（实时搜索）
        """
        # 获取搜索关键字
        keyword = self.search_var.get().strip()
        
        # 过滤并显示数据
        self._filter_and_display_data(keyword)
    
    def _on_clear_button_click(self):
        """
        清空按钮点击事件（清空搜索框并显示全量数据）
        """
        # 清空搜索框
        self.search_var.set("")
        
        # 直接显示全量数据（不重新加载数据库）
        self._filter_and_display_data("")
    
    def _filter_and_display_data(self, keyword: str = ""):
        """
        根据关键字过滤并显示数据
        
        Args:
            keyword: 搜索关键字（空字符串表示显示全部数据）
        """
        # 清空当前显示
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 隐藏空状态提示（默认）
        if self.empty_state_label:
            self.empty_state_label.place_forget()
        
        # 如果没有数据，直接返回
        if not self.all_data:
            # 更新搜索结果计数（无数据）
            if self.search_count_label:
                self.search_count_label.config(text="")
            return
        
        # 过滤数据（性能优化：使用列索引常量和列表推导式）
        if keyword:
            # 转换为小写进行模糊匹配
            keyword_lower = keyword.lower()
            
            # 使用列表推导式提升性能
            filtered_data = [
                item for item in self.all_data
                if keyword_lower in str(item['values'][COL_OBJ_NICK]).lower() 
                or keyword_lower in str(item['values'][COL_OBJ_TOTAL_NUMBER]).lower()
            ]
        else:
            # 无关键字，显示全部数据
            filtered_data = self.all_data
        
        # 更新搜索结果计数
        if self.search_count_label:
            if keyword:
                # 有搜索关键字时显示计数
                self.search_count_label.config(text=f"搜索到 {len(filtered_data)} 条数据")
            else:
                # 无搜索关键字时隐藏计数
                self.search_count_label.config(text="")
        
        # 显示过滤后的数据
        if filtered_data:
            for index, item in enumerate(filtered_data, start=1):
                # 添加行号
                values_with_row_num = (index,) + item['values']
                
                # 性能优化：预计算奇偶性，减少重复判断
                is_odd = index % 2 == 1
                original_tag = item['tag']
                
                # 根据原始标签类型和奇偶性分配新标签
                if 'deleted' in original_tag:
                    new_tag = "deleted_odd" if is_odd else "deleted_even"
                elif 'warning' in original_tag:
                    new_tag = "warning_odd" if is_odd else "warning_even"
                elif 'txt_file' in original_tag:
                    new_tag = "txt_file_odd" if is_odd else "txt_file_even"
                else:
                    new_tag = "oddrow" if is_odd else "evenrow"
                
                self.tree.insert("", tk.END, values=values_with_row_num, tags=(new_tag,))
        else:
            # 无匹配数据，显示居中的空状态提示
            if self.empty_state_label:
                self.empty_state_label.place(relx=0.5, rely=0.5, anchor="center")
    
    def _create_loading_widget(self):
        """创建Loading加载动画组件(现代化进度条设计)"""
        # 创建半透明遮罩层(浅灰色背景,柔和不刺眼)
        self.loading_frame = tk.Frame(
            self.root,
            bg="#f5f6fa",  # 与主窗口背景一致
            bd=0,
            highlightthickness=0
        )
        
        # 创建中央容器
        container = tk.Frame(
            self.loading_frame,
            bg="#f5f6fa",
            bd=0,
            highlightthickness=0
        )
        container.place(relx=0.5, rely=0.5, anchor="center")
        
        # 加载文字(在进度条上方)
        loading_label = tk.Label(
            container,
            text="正在加载数据...",
            font=("微软雅黑", 13),
            fg="#2c3e50",
            bg="#f5f6fa"
        )
        loading_label.pack(pady=(0, 20))
        
        # 创建Canvas用于绘制进度条
        self.loading_canvas = tk.Canvas(
            container,
            width=400,
            height=6,
            bg="#f5f6fa",
            bd=0,
            highlightthickness=0
        )
        self.loading_canvas.pack()
        
        # 绘制进度条背景(浅黄色)
        self.loading_canvas.create_rectangle(
            0, 0, 400, 6,
            fill="#FEF5E7",
            outline="",
            tags="bg"
        )
        
        # 绘制进度条(橙黄色,醒目,初始位置在最左侧)
        self.loading_bar_id = self.loading_canvas.create_rectangle(
            0, 0, 100, 6,
            fill="#f39c12",
            outline="",
            tags="bar"
        )
    
    def _show_loading(self):
        """显示Loading动画"""
        if self.loading_frame:
            # 显示遮罩层,覆盖整个窗口
            self.loading_frame.place(x=0, y=0, relwidth=1, relheight=1)
            # 强制刷新UI,确保Canvas完全渲染
            self.root.update_idletasks()
            # 启动进度条滑动动画
            self._animate_loading()
    
    def _hide_loading(self):
        """隐藏Loading动画"""
        # 停止动画
        if self.loading_animation_id is not None:
            self.root.after_cancel(self.loading_animation_id)
            self.loading_animation_id = None
        
        # 隐藏遮罩层
        if self.loading_frame:
            self.loading_frame.place_forget()
        
        # 重置位置
        self.loading_position = 0
        self.loading_direction = 1
    
    def _animate_loading(self):
        """Loading进度条滑动动画(快速左右滑动)"""
        if self.loading_canvas and self.loading_bar_id:
            # 更新进度条位置(大幅加快速度到20像素/次)
            self.loading_position += self.loading_direction * 20
            
            # 边界检测,到达边界时反向
            if self.loading_position >= 300:  # 右边界(400 - 100 = 300)
                self.loading_position = 300
                self.loading_direction = -1
            elif self.loading_position <= 0:  # 左边界
                self.loading_position = 0
                self.loading_direction = 1
            
            # 更新进度条位置
            self.loading_canvas.coords(
                self.loading_bar_id,
                self.loading_position, 0,
                self.loading_position + 100, 6
            )
            
            # 继续动画(每15毫秒更新一次,更快更流畅)
            self.loading_animation_id = self.root.after(15, self._animate_loading)
    
    def load_data(self):
        """加载数据库数据(入口方法 - 使用线程异步加载)"""
        
        # 重置搜索框
        self.search_var.set("")
        
        # 隐藏空状态提示
        if self.empty_state_label:
            self.empty_state_label.place_forget()
        
        # 清空搜索结果计数
        if self.search_count_label:
            self.search_count_label.config(text="")
        
        # 清空现有数据(在主线程中操作UI)
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 显示Loading动画
        self._show_loading()
        
        # 更新状态
        self.status_label.config(text="正在连接数据库...", fg="#3498db")
        
        # 在子线程中执行数据加载
        thread = threading.Thread(target=self._load_data_in_thread)
        thread.daemon = True  # 守护线程,主程序退出时自动结束
        thread.start()
    
    def _load_data_in_thread(self):
        """在子线程中执行数据加载"""
        
        # ========== 性能日志：开始计时 ==========
        start_time = time.time()
        print(f"\n{'='*60}")
        print(f"[性能日志] 开始加载数据")
        print(f"{'='*60}")
        
        # ========== 新增：检测主数据库是否存在（仅首次启动时提示） ==========
        if self.is_first_load and not os.path.exists(PRIMARY_DB):
            # 主数据库不存在，标记需要显示警告（在数据加载完成后显示）
            self.need_show_db_warning = True
        
        # 标记已完成首次加载
        self.is_first_load = False
        
        # 获取数据库连接(在子线程中执行)
        step_start = time.time()
        conn, message = get_database_connection()
        print(f"[性能日志] 1. 获取数据库连接: {(time.time() - step_start)*1000:.2f}ms")
        
        if not conn:
            # 连接失败 - 调度主线程更新UI
            error_msg = message
            self.root.after(0, lambda: self._on_load_error("连接失败", f"无法连接到数据库:\n\n{error_msg}"))
            return
        
        self.conn = conn
        
        # 提取数据库路径(从消息中)
        if ":" in message:
            self.current_db_path = message.split(":", 1)[1].strip()
        
        # ========== 新增：异步备份超微数据库（性能优化） ==========
        # 如果连接成功且使用的是主数据库，则异步备份
        if self.current_db_path:
            # 只有当前连接的是主数据库时才备份（使用模块级常量）
            if self.current_db_path == PRIMARY_DB:
                # 异步备份到本地，不阻塞数据加载
                backup_thread = threading.Thread(
                    target=backup_database,
                    args=(PRIMARY_DB, BACKUP_DB),
                    daemon=True
                )
                backup_thread.start()
                print(f"[性能日志] 2. 异步备份超微数据库到本地（已启动后台线程）")
                
                # 异步备份到网络路径（带时间标签）
                network_backup_thread = threading.Thread(
                    target=backup_to_network,
                    args=(PRIMARY_DB,),
                    daemon=True
                )
                network_backup_thread.start()
                print(f"[性能日志] 3. 异步备份超微数据库到网络（已启动后台线程）")
        
        # 生成原有数据库精简路径（最后4级）
        main_db_short_path = ""
        if self.current_db_path:
            main_db_short_path = get_short_path(self.current_db_path, num_parts=4)
        
        # ========== 新增：优化 contact.db 连接逻辑（性能优化） ==========
        step_start = time.time()
        try:
            # 获取 contact.db 网络路径
            network_db_path = get_contact_db_path()
            
            # 如果网络路径存在，先备份到桌面，然后访问桌面文件
            if network_db_path.startswith(CONTACT_NETWORK_BASE) and os.path.exists(network_db_path):
                backup_start = time.time()
                backup_success = backup_database(network_db_path, CONTACT_LOCAL_BACKUP)
                print(f"[性能日志] 4. 备份 contact.db 到桌面: {(time.time() - backup_start)*1000:.2f}ms")
                
                # 使用本地备份（性能优化：从本地访问更快）
                self.contact_db_path = CONTACT_LOCAL_BACKUP
            else:
                # 网络路径不存在，直接使用桌面路径
                self.contact_db_path = CONTACT_LOCAL_BACKUP
                print(f"[性能日志] 4. 网络路径不存在，直接使用桌面 contact.db")
            
            # 连接本地 contact.db
            if os.path.exists(self.contact_db_path):
                connect_start = time.time()
                self.contact_conn = sqlite3.connect(self.contact_db_path)
                # 验证连接可用性
                cursor = self.contact_conn.cursor()
                cursor.execute("SELECT 1")
                cursor.close()
                print(f"[性能日志] 5. 连接本地 contact.db: {(time.time() - connect_start)*1000:.2f}ms")
            else:
                self.contact_conn = None
                print(f"[性能日志] 5. 本地 contact.db 不存在")
                
        except Exception as e:
            # 连接失败，静默处理
            self.contact_conn = None
            print(f"[备份日志] contact.db 连接失败: {e}")
            pass
        
        # 更新状态 - 调度主线程更新UI
        self.root.after(0, lambda: self.status_label.config(text="正在读取数据...", fg="#3498db"))
        
        # 读取数据(在子线程中执行)
        try:
            
            cursor = conn.cursor()
            
            # ========== 预加载：构建微信ID到昵称、微信号、添加时间和删除状态的映射表（性能优化） ==========
            step_start = time.time()
            wxid_info = {}
            try:
                # 性能优化：添加 WHERE 条件，过滤空 WxID
                cursor.execute("SELECT WxID, Nick, Number, AddTime, Deleted FROM ContactConfigTable WHERE WxID IS NOT NULL")
                
                # 性能优化：使用字典推导式批量构建映射
                wxid_info = {
                    row[0]: {
                        'nick': row[1] or "",
                        'number': row[2] or "",
                        'addtime': row[3] or 0,
                        'deleted': row[4] or 0
                    }
                    for row in cursor.fetchall()
                }
                print(f"[性能日志] 6. 预加载 ContactConfigTable ({len(wxid_info)} 条): {(time.time() - step_start)*1000:.2f}ms")
            except sqlite3.OperationalError:
                # ContactConfigTable 表不存在,使用空映射
                print("警告: ContactConfigTable 表不存在,昵称、微信号、添加时间和删除状态列将为空")
                wxid_info = {}
            
            # ========== 预加载：构建 contact.db 的 username 到完整信息的映射表（性能优化：流式处理） ==========
            step_start = time.time()
            contact_info_map = {}
            valid_count = 0
            invalid_count = 0
            if self.contact_conn:
                try:
                    contact_cursor = self.contact_conn.cursor()
                    # 性能优化：只查询需要的列，减少数据传输
                    contact_cursor.execute("SELECT username, remark, nick_name, alias FROM contact WHERE username IS NOT NULL")
                    
                    # 性能优化：使用流式处理（fetchmany），减少内存峰值
                    batch_size = 1000
                    while True:
                        rows = contact_cursor.fetchmany(batch_size)
                        if not rows:
                            break
                        
                        for row in rows:
                            username = row[0]
                            remark = row[1] if row[1] else ""
                            nick_name = row[2] if row[2] else ""
                            alias = row[3] if row[3] else ""
                            
                            # 验证 remark 格式（¿¿¿ + 6个数字 + - + 其他内容）
                            if self.validate_remark_format(remark):
                                contact_info_map[username] = {
                                    'remark': remark,
                                    'nick_name': nick_name,
                                    'alias': alias
                                }
                                valid_count += 1
                            else:
                                contact_info_map[username] = {
                                    'remark': "",
                                    'nick_name': nick_name,
                                    'alias': alias
                                }
                                invalid_count += 1
                    
                    contact_cursor.close()
                    print(f"[性能日志] 7. 预加载 contact 表 (总计: {len(contact_info_map)}, 有效: {valid_count}, 无效: {invalid_count}): {(time.time() - step_start)*1000:.2f}ms")
                except (sqlite3.OperationalError, Exception) as e:
                    # contact 表不存在或查询失败,使用空映射
                    contact_info_map = {}
                    print(f"[性能日志] 7. 预加载 contact 表失败: {e}")
            
            # ========== 预加载：构建售前通讯录.txt的微信ID到添加时间的映射表（性能优化：流式读取） ==========
            step_start = time.time()
            txt_time_map = {}  # {微信ID: 添加时间}
            txt_file_path = get_resource_path("售前通讯录.txt")
            try:
                if os.path.exists(txt_file_path):
                    # 性能优化：使用生成器逐行读取，避免一次性加载整个文件
                    with open(txt_file_path, 'r', encoding='utf-8') as f:
                        # 跳过表头
                        next(f, None)
                        
                        for line in f:
                            # 去除首尾空白字符
                            line = line.rstrip('\r\n')
                            
                            # 跳过完全空白的行
                            if not line.strip():
                                continue
                            
                            # 按制表符分割
                            parts = line.split('\t')
                            
                            # 提取微信ID（第2列）和添加时间（第5列）
                            if len(parts) >= 5:
                                wxid = parts[1].strip()
                                time_str = parts[4].strip()
                                
                                # 转换时间格式（处理"2025年3月"格式）
                                if wxid and time_str:
                                    converted_time = self.convert_time_format(time_str)
                                    if converted_time:
                                        txt_time_map[wxid] = converted_time
                    
                    print(f"[性能日志] 8. 预加载售前通讯录.txt时间映射 ({len(txt_time_map)} 条): {(time.time() - step_start)*1000:.2f}ms")
                else:
                    print(f"[性能日志] 8. 售前通讯录.txt 不存在，跳过时间映射")
            except Exception as e:
                print(f"[性能日志] 8. 预加载售前通讯录.txt时间映射失败: {e}")
                txt_time_map = {}
            
            # ========== 第一部分：读取 AutoAgreeAddFriendTask 表 ==========
            step_start = time.time()
            cursor.execute("SELECT wxid, time, content FROM AutoAgreeAddFriendTask")
            rows = cursor.fetchall()
            print(f"[性能日志] 9. 读取 AutoAgreeAddFriendTask ({len(rows)} 条): {(time.time() - step_start)*1000:.2f}ms")
            
            # 准备要插入的数据列表(在子线程中处理数据)
            data_to_insert = []
            
            # 第一步：预处理数据，构建每个wxid的最早记录字典
            step_start = time.time()
            wxid_earliest = {}  # {wxid: {'nickname': str, 'username': str, 'time': int}}
            
            for row in rows:
                wxid_value = row[0] if row[0] is not None else ""
                time_value = row[1] if row[1] is not None else 0
                content_value = row[2] if row[2] is not None else ""
                
                # 解析XML提取用户名和昵称
                username, nickname = self.parse_xml_content(content_value)
                
                # 如果该wxid还没有记录，或者当前时间更早，则更新
                if wxid_value:
                    if wxid_value not in wxid_earliest:
                        wxid_earliest[wxid_value] = {
                            'nickname': nickname,
                            'username': username,
                            'time': time_value
                        }
                    else:
                        # 比较时间，保留最早的记录
                        if time_value < wxid_earliest[wxid_value]['time']:
                            wxid_earliest[wxid_value] = {
                                'nickname': nickname,
                                'username': username,
                                'time': time_value
                            }
            
            print(f"[性能日志] 10. 预处理数据（构建最早记录）: {(time.time() - step_start)*1000:.2f}ms")
            
            # 第二步：插入数据到表格，并标记不一致的数据
            step_start = time.time()
            for index, row in enumerate(rows):
                # 提取原始数据
                wxid_value = row[0] if row[0] is not None else ""
                time_value = row[1] if row[1] is not None else 0
                content_value = row[2] if row[2] is not None else ""
                
                # 转换时间戳
                formatted_time = self.convert_timestamp(time_value)
                
                # 备用时间源：如果时间为空，从售前通讯录.txt映射表查询
                if not formatted_time and wxid_value and wxid_value in txt_time_map:
                    formatted_time = txt_time_map[wxid_value]
                
                # 解析XML提取用户名和昵称
                username, nickname = self.parse_xml_content(content_value)
                
                # 检查是否需要标记（空值统一标记，非空值比较一致性）
                needs_warning = False
                display_nickname = nickname
                display_username = username
                
                # 处理昵称：空值直接标记
                if not nickname or nickname.strip() == "":
                    display_nickname = "【需二次确认】"
                    needs_warning = True
                elif wxid_value and wxid_value in wxid_earliest:
                    # 非空值：如果不是最早记录且与最早记录不同，则标记
                    earliest = wxid_earliest[wxid_value]
                    if time_value != earliest['time'] and nickname != earliest['nickname']:
                        display_nickname = f"{nickname} 【需二次确认】"
                        needs_warning = True
                
                # 处理用户名：空值直接标记
                if not username or username.strip() == "":
                    display_username = "【需二次确认】"
                    needs_warning = True
                elif wxid_value and wxid_value in wxid_earliest:
                    # 非空值：如果不是最早记录且与最早记录不同，则标记
                    earliest = wxid_earliest[wxid_value]
                    if time_value != earliest['time'] and username != earliest['username']:
                        display_username = f"{username} 【需二次确认】"
                        needs_warning = True
                
                # 查询昵称、微信号和删除状态(从预加载的映射表中获取)
                # 如果wxid_value不在映射表中，视为已删除
                if wxid_value and wxid_value in wxid_info:
                    info = wxid_info[wxid_value]
                    nick = clean_newlines(info['nick'])  # 清理对象昵称中的换行符
                    number = info['number']
                    deleted_value = info.get('deleted', 0)
                else:
                    # WxID不存在于ContactConfigTable，视为已删除
                    nick = ""
                    number = ""
                    deleted_value = 1  # 标记为已删除
                
                # 计算"对象(总微信号)"：如果number非空则用number，否则用wxid_value
                total_number = number if number and number.strip() else wxid_value
                
                # 判断是否删除：如果deleted==1，显示"已删除"(红色)
                is_delete_text = "已删除" if deleted_value == 1 else ""
                
                # 查询"对象(内部备注)"：使用对象微信ID去contact.db查询
                obj_internal_note = ""
                if wxid_value:
                    contact_data = contact_info_map.get(wxid_value, None)
                    if contact_data and contact_data['remark']:
                        # remark已经在预加载时验证过格式，这里直接使用
                        obj_internal_note = contact_data['remark']
                
                # 三级数据优先级：来源相关列的数据填充
                if not username or username.strip() == "":
                    # 来源微信ID为空，所有来源列填充【需二次确认】
                    display_nickname = "【需二次确认】"
                    source_number = "【需二次确认】"
                    source_total_number = "【需二次确认】"
                    internal_note = ""
                else:
                    # 第一优先级：ContactConfigTable
                    source_info = wxid_info.get(username, None)
                    if source_info:
                        # 使用ContactConfigTable数据
                        display_nickname = clean_newlines(source_info['nick'])  # 清理来源昵称中的换行符
                        source_number = source_info['number']
                        # 查询内部备注
                        contact_data = contact_info_map.get(username, None)
                        internal_note = contact_data['remark'] if contact_data else ""
                    else:
                        # 第二优先级：contact.db
                        contact_data = contact_info_map.get(username, None)
                        if contact_data:
                            # 使用contact.db数据
                            display_nickname = clean_newlines(contact_data['nick_name'])  # 清理来源昵称中的换行符
                            source_number = contact_data['alias']
                            internal_note = contact_data['remark']
                        else:
                            # 第三优先级：数据库数据填空串
                            display_nickname = ""
                            source_number = ""
                            internal_note = ""
                    
                    # 计算"来源(总微信号)"：如果source_number非空则用source_number，否则用username
                    source_total_number = source_number if source_number and source_number.strip() else username
                    
                    # 【需二次确认】标记传播逻辑：如果"来源微信ID"包含标记，则传播到其他来源列
                    if "【需二次确认】" in display_username:
                        # 如果"来源昵称"非空，添加标记
                        if display_nickname and display_nickname.strip() and display_nickname != "【需二次确认】":
                            display_nickname = f"{display_nickname} 【需二次确认】"
                        # 如果"来源微信号"非空，添加标记
                        if source_number and source_number.strip() and source_number != "【需二次确认】":
                            source_number = f"{source_number} 【需二次确认】"
                        # "来源总微信号"始终添加标记（无论是否为空）
                        if source_total_number != "【需二次确认】":
                            source_total_number = f"{source_total_number} 【需二次确认】"
                
                # 选择标签（交替行颜色 + 警告样式 + 已删除样式）
                if deleted_value == 1:
                    # 已删除的联系人用红色标记
                    tag = "deleted_even" if index % 2 == 0 else "deleted_odd"
                elif needs_warning:
                    # 需要二次确认的数据用红色标记
                    tag = "warning_even" if index % 2 == 0 else "warning_odd"
                else:
                    # 正常数据用交替行颜色
                    tag = "evenrow" if index % 2 == 0 else "oddrow"
                
                # 添加到待插入列表（新列顺序：对象相关列在前，来源相关列在后）
                # 列顺序：对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、来源总微信号、来源内部备注
                data_to_insert.append({
                    'values': (nick, wxid_value, number, total_number, formatted_time, obj_internal_note, is_delete_text, display_nickname, display_username, source_number, source_total_number, internal_note),
                    'tag': tag
                })
            
            print(f"[性能日志] 11. 处理 AutoAgreeAddFriendTask 数据: {(time.time() - step_start)*1000:.2f}ms")
            
            # ========== 第二部分：读取 ContactAuthMsgTable 表并追加 ==========
            step_start = time.time()
            # 构建已存在数据的去重集合（基于对象微信ID和来源微信ID）
            existing_data_set = set()
            for item in data_to_insert:
                values = item['values']
                # 使用列索引常量，提升代码可读性
                if len(values) > COL_SOURCE_WXID:
                    # 使用优化的清理函数
                    col_wxid = str(values[COL_OBJ_WXID]).strip()  # 对象(微信ID)
                    col_source_username = clean_confirmation_mark(str(values[COL_SOURCE_WXID]))  # 来源(微信ID)
                    existing_data_set.add((col_wxid, col_source_username))
            
            # 查询 ContactAuthMsgTable 表
            try:
                cursor.execute("SELECT WxID, Xml FROM ContactAuthMsgTable")
                contact_rows = cursor.fetchall()
                print(f"[性能日志] 12. 读取 ContactAuthMsgTable ({len(contact_rows)} 条): {(time.time() - step_start)*1000:.2f}ms")
                
                step_start = time.time()
                new_data_count = 0
                current_total = len(data_to_insert)
                
                for row in contact_rows:
                    wxid_value = row[0] if row[0] is not None else ""
                    xml_value = row[1] if row[1] is not None else ""
                    
                    # 解析XML提取用户名和昵称（复用现有逻辑）
                    username, nickname = self.parse_xml_content(xml_value)
                    
                    # 去重检查：只匹配对象微信ID和来源微信ID
                    data_key = (wxid_value.strip(), username.strip())
                    if data_key in existing_data_set:
                        continue
                    
                    # 添加到去重集合
                    existing_data_set.add(data_key)
                    
                    # 查询昵称、微信号、添加时间和删除状态(从预加载的映射表中获取)
                    # 如果wxid_value不在映射表中，视为已删除
                    if wxid_value and wxid_value in wxid_info:
                        info = wxid_info[wxid_value]
                        nick = clean_newlines(info['nick'])  # 清理对象昵称中的换行符
                        number = info['number']
                        addtime = info.get('addtime', 0)
                        deleted_value = info.get('deleted', 0)
                    else:
                        # WxID不存在于ContactConfigTable，视为已删除
                        nick = ""
                        number = ""
                        addtime = 0
                        deleted_value = 1  # 标记为已删除
                    
                    # 计算"对象(总微信号)"：如果number非空则用number，否则用wxid_value
                    total_number = number if number and number.strip() else wxid_value
                    
                    # 判断是否删除：如果deleted==1，显示"已删除"(红色)
                    is_delete_text = "已删除" if deleted_value == 1 else ""
                    
                    # 第5列添加时间：从ContactConfigTable的AddTime字段转换(.NET Ticks格式)
                    formatted_time = self.convert_dotnet_ticks(addtime) if addtime else ""
                    
                    # 备用时间源：如果时间为空，从售前通讯录.txt映射表查询
                    if not formatted_time and wxid_value and wxid_value in txt_time_map:
                        formatted_time = txt_time_map[wxid_value]
                    
                    # 查询"对象(内部备注)"：使用对象微信ID去contact.db查询
                    obj_internal_note = ""
                    if wxid_value:
                        contact_data = contact_info_map.get(wxid_value, None)
                        if contact_data and contact_data['remark']:
                            # remark已经在预加载时验证过格式，这里直接使用
                            obj_internal_note = contact_data['remark']
                    
                    # 三级数据优先级：来源相关列的数据填充
                    if not username or username.strip() == "":
                        # 来源微信ID为空，所有来源列填充【需二次确认】
                        nickname = "【需二次确认】"
                        username = "【需二次确认】"
                        source_number = "【需二次确认】"
                        source_total_number = "【需二次确认】"
                        internal_note = ""
                    else:
                        # 第一优先级：ContactConfigTable
                        source_info = wxid_info.get(username, None)
                        if source_info:
                            # 使用ContactConfigTable数据
                            nickname = clean_newlines(source_info['nick'])  # 清理来源昵称中的换行符
                            source_number = source_info['number']
                            # 查询内部备注
                            contact_data = contact_info_map.get(username, None)
                            internal_note = contact_data['remark'] if contact_data else ""
                        else:
                            # 第二优先级：contact.db
                            contact_data = contact_info_map.get(username, None)
                            if contact_data:
                                # 使用contact.db数据
                                nickname = clean_newlines(contact_data['nick_name'])  # 清理来源昵称中的换行符
                                source_number = contact_data['alias']
                                internal_note = contact_data['remark']
                            else:
                                # 第三优先级：数据库数据填空串
                                nickname = ""
                                source_number = ""
                                internal_note = ""
                        
                        # 计算"来源(总微信号)"：如果source_number非空则用source_number，否则用username
                        source_total_number = source_number if source_number and source_number.strip() else username
                        
                        # 【需二次确认】标记传播逻辑：如果"来源微信ID"包含标记，则传播到其他来源列
                        if "【需二次确认】" in username:
                            # 如果"来源昵称"非空，添加标记
                            if nickname and nickname.strip() and nickname != "【需二次确认】":
                                nickname = f"{nickname} 【需二次确认】"
                            # 如果"来源微信号"非空，添加标记
                            if source_number and source_number.strip() and source_number != "【需二次确认】":
                                source_number = f"{source_number} 【需二次确认】"
                            # "来源总微信号"始终添加标记（无论是否为空）
                            if source_total_number != "【需二次确认】":
                                source_total_number = f"{source_total_number} 【需二次确认】"
                    
                    # 选择标签（已删除用红色，正常数据用黑色）
                    if deleted_value == 1:
                        # 已删除的联系人用红色标记
                        tag = "deleted_even" if (current_total + new_data_count) % 2 == 0 else "deleted_odd"
                    else:
                        # 正常数据用黑色（交替行颜色）
                        tag = "evenrow" if (current_total + new_data_count) % 2 == 0 else "oddrow"
                    
                    # 添加到待插入列表（新列顺序：对象相关列在前，来源相关列在后）
                    # 列顺序：对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、来源总微信号、来源内部备注
                    data_to_insert.append({
                        'values': (nick, wxid_value, number, total_number, formatted_time, obj_internal_note, is_delete_text, nickname, username, source_number, source_total_number, internal_note),
                        'tag': tag
                    })
                    new_data_count += 1
                
                print(f"[性能日志] 13. 处理 ContactAuthMsgTable 数据: {(time.time() - step_start)*1000:.2f}ms")
                
                # 准备状态消息
                total_count = len(rows) + new_data_count
                status_msg = f"✓ 加载完成，共 {total_count} 条记录（AutoAgreeAddFriendTask: {len(rows)}，ContactAuthMsgTable: {new_data_count}）"
                
            except sqlite3.OperationalError as e:
                # ContactAuthMsgTable 表可能不存在，仅记录警告，不中断流程
                print(f"警告: 无法读取 ContactAuthMsgTable 表: {e}")
                status_msg = f"✓ 加载完成，共 {len(rows)} 条记录（仅 AutoAgreeAddFriendTask）"
            
            cursor.close()
            
            # ========== 第三部分：读取售前通讯录.txt并追加 ==========
            step_start = time.time()
            txt_data_count = 0
            skipped_lines = []  # 记录跳过的行
            
            try:
                if os.path.exists(txt_file_path):
                    with open(txt_file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    print(f"[调试] 售前通讯录.txt 总行数: {len(lines)}，数据行数（不含表头）: {len(lines) - 1}")
                    
                    # 跳过表头（第一行）
                    current_total = len(data_to_insert)
                    
                    for line_idx, line in enumerate(lines[1:], start=1):
                        # 去除首尾空白字符（但保留制表符）
                        line = line.rstrip('\r\n')
                        
                        # 跳过完全空白的行
                        if not line.strip():
                            skipped_lines.append((line_idx + 1, "空行"))
                            continue
                        
                        # 按制表符分割
                        parts = line.split('\t')
                        
                        # 增强容错：如果列数不足12列，用空字符串填充
                        while len(parts) < 12:
                            parts.append('')
                        
                        # 提取各列数据（安全提取，防止索引越界）- 新列顺序（12列）
                        # 列顺序：对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、来源总微信号、来源内部备注
                        nick_from_file = parts[0].strip() if len(parts) > 0 else ""
                        wxid_value = parts[1].strip() if len(parts) > 1 else ""
                        number_from_file = parts[2].strip() if len(parts) > 2 else ""
                        total_number_from_file = parts[3].strip() if len(parts) > 3 else ""
                        formatted_time = parts[4].strip() if len(parts) > 4 else ""
                        obj_internal_note_from_file = parts[5].strip() if len(parts) > 5 else ""
                        is_delete_symbol = parts[6].strip() if len(parts) > 6 else ""
                        source_nickname = parts[7].strip() if len(parts) > 7 else ""
                        source_username = parts[8].strip() if len(parts) > 8 else ""
                        source_number = parts[9].strip() if len(parts) > 9 else ""
                        source_total_number = parts[10].strip() if len(parts) > 10 else ""
                        source_internal_note = parts[11].strip() if len(parts) > 11 else ""
                        
                        # 先转换文本文件中的时间格式（作为备用）
                        formatted_time_from_file = self.convert_time_format(formatted_time)
                        
                        # 统一删除状态判断：用wxid_value匹配ContactConfigTable
                        if wxid_value and wxid_value in wxid_info:
                            # 匹配成功，使用数据库数据
                            info = wxid_info[wxid_value]
                            nick = clean_newlines(info['nick'])  # 使用数据库Nick，清理换行符
                            number = info['number']  # 使用数据库Number
                            addtime = info.get('addtime', 0)
                            deleted_value = info.get('deleted', 0)
                            
                            # 时间：优先使用数据库AddTime，为空则使用文本文件时间
                            if addtime:
                                formatted_time = self.convert_dotnet_ticks(addtime)
                            else:
                                formatted_time = formatted_time_from_file
                        else:
                            # 匹配失败，标记为已删除
                            nick = clean_newlines(nick_from_file)  # 使用文本文件Nick，清理换行符
                            number = number_from_file
                            formatted_time = formatted_time_from_file
                            deleted_value = 1  # 标记为已删除
                        
                        # 计算"对象(总微信号)"：如果number非空则用number，否则用wxid_value
                        total_number = number if number and number.strip() else wxid_value
                        
                        # 判断是否删除：如果deleted==1，显示"已删除"
                        is_delete_text = "已删除" if deleted_value == 1 else ""
                        
                        # 查询"对象(内部备注)"：优先使用contact.db数据，否则使用文本文件数据
                        obj_internal_note = ""
                        if wxid_value:
                            contact_data = contact_info_map.get(wxid_value, None)
                            if contact_data and contact_data['remark']:
                                # 使用contact.db数据（已验证格式）
                                obj_internal_note = contact_data['remark']
                            else:
                                # 使用文本文件数据（需要验证格式）
                                if obj_internal_note_from_file and self.validate_remark_format(obj_internal_note_from_file):
                                    obj_internal_note = obj_internal_note_from_file
                        
                        # 三级数据优先级：来源相关列的数据填充（售前通讯录.txt）
                        if not source_username or source_username.strip() == "":
                            # 来源微信ID为空，使用文本文件数据
                            final_source_nickname = clean_newlines(source_nickname)  # 清理来源昵称中的换行符
                            final_source_number = source_number
                            final_source_total_number = source_total_number
                            final_source_internal_note = source_internal_note
                        else:
                            # 第一优先级：ContactConfigTable
                            source_info = wxid_info.get(source_username, None)
                            if source_info:
                                # 使用ContactConfigTable数据
                                final_source_nickname = clean_newlines(source_info['nick'])  # 清理来源昵称中的换行符
                                final_source_number = source_info['number']
                                # 查询内部备注
                                contact_data = contact_info_map.get(source_username, None)
                                final_source_internal_note = contact_data['remark'] if contact_data else ""
                            else:
                                # 第二优先级：contact.db
                                contact_data = contact_info_map.get(source_username, None)
                                if contact_data:
                                    # 使用contact.db数据
                                    final_source_nickname = clean_newlines(contact_data['nick_name'])  # 清理来源昵称中的换行符
                                    final_source_number = contact_data['alias']
                                    final_source_internal_note = contact_data['remark']
                                else:
                                    # 第三优先级：使用文本文件原始数据
                                    final_source_nickname = clean_newlines(source_nickname)  # 清理来源昵称中的换行符
                                    final_source_number = source_number
                                    final_source_internal_note = source_internal_note
                            
                            # 计算"来源(总微信号)"：如果final_source_number非空则用final_source_number，否则用source_username
                            final_source_total_number = final_source_number if final_source_number and final_source_number.strip() else source_username
                        
                        # 选择标签（已删除用红色，正常数据用黑色）
                        if deleted_value == 1:
                            tag = "deleted_even" if (current_total + txt_data_count) % 2 == 0 else "deleted_odd"
                        else:
                            tag = "txt_file_even" if (current_total + txt_data_count) % 2 == 0 else "txt_file_odd"
                        
                        # 添加到待插入列表（新列顺序：对象相关列在前，来源相关列在后）
                        # 列顺序：对象昵称、对象微信ID、对象微信号、对象总微信号、对象添加时间、对象内部备注、对象是否删除、来源昵称、来源微信ID、来源微信号、来源总微信号、来源内部备注
                        data_to_insert.append({
                            'values': (nick, wxid_value, number, total_number, formatted_time, obj_internal_note, is_delete_text, final_source_nickname, source_username, final_source_number, final_source_total_number, final_source_internal_note),
                            'tag': tag
                        })
                        txt_data_count += 1
                    
                    # 输出跳过的行信息
                    if skipped_lines:
                        print(f"[调试] 跳过 {len(skipped_lines)} 行:")
                        for line_num, reason in skipped_lines[:10]:  # 只显示前10条
                            print(f"  - 第 {line_num} 行: {reason}")
                        if len(skipped_lines) > 10:
                            print(f"  - ... 还有 {len(skipped_lines) - 10} 行")
                    
                    print(f"[性能日志] 14. 读取售前通讯录.txt ({txt_data_count} 条): {(time.time() - step_start)*1000:.2f}ms")
                    
                    # 更新状态消息
                    total_count = len(rows) + new_data_count + txt_data_count
                    status_msg = f"✓ 加载完成，共 {total_count} 条记录（AutoAgreeAddFriendTask: {len(rows)}，ContactAuthMsgTable: {new_data_count}，售前通讯录: {txt_data_count}）"
                else:
                    print(f"[性能日志] 14. 售前通讯录.txt 不存在，跳过")
            
            except Exception as e:
                # 文本文件读取失败，仅记录警告，不中断流程
                print(f"警告: 无法读取售前通讯录.txt: {e}")
                import traceback
                traceback.print_exc()
                pass
            
            # 生成 contact.db 精简路径（最后4级）
            contact_db_short_path = ""
            if self.contact_db_path:
                contact_db_short_path = get_short_path(self.contact_db_path, num_parts=4)
            
            # 性能日志：总耗时
            total_time = time.time() - start_time
            print(f"\n[性能日志] 15. 数据处理总耗时: {total_time*1000:.2f}ms ({total_time:.2f}秒)")
            print(f"{'='*60}\n")
            
            # 调度主线程更新UI（传递两个数据库路径）
            ui_start = time.time()
            self.root.after(0, lambda: self._update_ui_with_data(data_to_insert, status_msg, main_db_short_path, contact_db_short_path, ui_start))
            
        except sqlite3.OperationalError as e:
            # 表不存在或SQL错误 - 调度主线程更新UI
            error_msg = f"查询失败: {e}"
            detail_msg = f"无法读取表数据:\n\n{error_msg}\n\n可能原因:\n- 表 'AutoAgreeAddFriendTask' 不存在\n- 列名不正确"
            self.root.after(0, lambda: self._on_load_error("查询错误", detail_msg))
        
        except Exception as e:
            # 其他错误 - 调度主线程更新UI
            error_msg = f"未知错误: {type(e).__name__}: {e}"
            self.root.after(0, lambda: self._on_load_error("错误", error_msg))
        
        finally:
            # 关闭主数据库连接（contact.db 连接保留用于后续数据交叉）
            # 重要：确保数据库连接被正确关闭，释放资源
            try:
                if conn:
                    conn.close()
            except Exception:
                # 静默处理关闭异常，避免影响错误处理流程
                pass
    
    def _deduplicate_data(self, data_list: list) -> list:
        """
        对数据进行去重，保留添加时间最早的记录
        
        去重规则：
            - 第2列（对象微信ID）和第8列（来源微信ID）完全一致视为重复
            - 只对添加时间非空的数据进行去重
            - 保留添加时间最早的一条记录
            - 添加时间为空的数据不参与去重，全部保留
        
        Args:
            data_list: 待去重的数据列表，每项包含 'values' 和 'tag'
        
        Returns:
            去重后的数据列表
            
        性能:
            - 时间复杂度: O(n)，使用字典去重
            - 空间复杂度: O(n)，需要存储去重字典
            
        注意:
            - 清理【需二次确认】标记后再进行去重判断
            - 时间字符串比较使用字典序（YYYY-MM-DD HH:MM:SS格式）
        """
        # 构建去重字典：key=(col2, col8), value=最早的记录
        unique_records = {}
        non_duplicate_records = []  # 不参与去重的记录（添加时间为空）
        
        for item in data_list:
            values = item['values']
            # 使用列索引常量，提升代码可读性和可维护性
            if len(values) > COL_SOURCE_WXID:
                col_wxid = str(values[COL_OBJ_WXID]).strip()  # 对象微信ID
                col_source_username = str(values[COL_SOURCE_WXID]).strip()  # 来源微信ID
                col_time = str(values[COL_OBJ_TIME]).strip()  # 添加时间
                
                # 使用优化的清理函数
                col_source_username_clean = clean_confirmation_mark(col_source_username)
                
                # 只对添加时间非空的数据进行去重
                if col_time and col_time != "":
                    key = (col_wxid, col_source_username_clean)
                    
                    if key not in unique_records:
                        # 第一次出现，直接保存
                        unique_records[key] = item
                    else:
                        # 比较添加时间，保留更早的
                        existing_time = str(unique_records[key]['values'][COL_OBJ_TIME]).strip()
                        if col_time < existing_time:  # 字符串比较（YYYY-MM-DD HH:MM:SS格式）
                            unique_records[key] = item
                else:
                    # 添加时间为空，不参与去重，直接保留
                    non_duplicate_records.append(item)
        
        # 合并去重后的数据和不参与去重的数据
        result = list(unique_records.values()) + non_duplicate_records
        
        return result
    
    def _update_ui_with_data(self, data_to_insert: list, status_msg: str, main_db_short_path: str = "", contact_db_short_path: str = "", ui_start: float = 0):
        """在主线程中更新UI(插入数据到表格)"""
        
        # 去重前的数据条数
        original_count = len(data_to_insert)
        
        # 执行全局去重
        dedup_start = time.time()
        data_to_insert = self._deduplicate_data(data_to_insert)
        dedup_count = len(data_to_insert)
        removed_count = original_count - dedup_count
        print(f"[性能日志] 16. 数据去重 (原始: {original_count}, 去重后: {dedup_count}, 移除: {removed_count}): {(time.time() - dedup_start)*1000:.2f}ms")
        
        # 保存完整数据集（用于搜索过滤）
        self.all_data = data_to_insert.copy()
        
        # 插入所有数据到表格（添加行号）
        insert_start = time.time()
        for index, item in enumerate(data_to_insert, start=1):
            # 在数据前添加行号
            values_with_row_num = (index,) + item['values']
            self.tree.insert("", tk.END, values=values_with_row_num, tags=(item['tag'],))
        
        print(f"[性能日志] 17. UI插入数据 ({dedup_count} 条): {(time.time() - insert_start)*1000:.2f}ms")
        if ui_start > 0:
            print(f"[性能日志] 18. UI更新总耗时: {(time.time() - ui_start)*1000:.2f}ms")
        
        # 更新状态消息，显示去重信息
        if removed_count > 0:
            # 使用正则表达式准确替换总条数
            import re
            status_msg = re.sub(
                r'共\s*\d+\s*条',
                f'去重前 {original_count} 条，去重后 {dedup_count} 条',
                status_msg
            )
        
        # 更新状态
        self.status_label.config(text=status_msg, fg="#2ecc71")
        
        # 更新原有数据库路径显示（精简路径）
        if main_db_short_path:
            self.main_db_label.config(text=f"超微数据库路径: {main_db_short_path}")
        else:
            self.main_db_label.config(text="")
        
        # 更新 contact.db 路径显示（精简路径）
        if contact_db_short_path:
            self.contact_db_label.config(text=f"内部数据库路径: {contact_db_short_path}")
        else:
            self.contact_db_label.config(text="内部数据库路径: 未连接")
        
        # 隐藏Loading动画
        self._hide_loading()
        
        # 性能优化：触发垃圾回收，释放临时对象占用的内存
        gc.collect()
        
        # ========== 新增：在数据加载完成后显示数据库警告 ==========
        if self.need_show_db_warning:
            self.need_show_db_warning = False
            # 延迟500毫秒显示Toast，确保UI完全加载完成
            self.root.after(500, lambda: self._show_warning_toast("请主动备份超微数据库", duration_ms=5000))
        
        # ========== 启动定时自动刷新（每5分钟） ==========
        self._start_auto_refresh()
    
    def _on_load_error(self, title: str, message: str):
        """在主线程中处理加载错误"""
        
        # 隐藏Loading
        self._hide_loading()
        
        # 更新状态
        self.status_label.config(text=message.split('\n')[0], fg="#e74c3c")
        
        # 显示错误对话框
        messagebox.showerror(title, message)
    
    def _start_auto_refresh(self):
        """
        启动定时自动刷新
        
        功能：
            - 每5分钟自动刷新一次数据
            - 在程序启动后或手动刷新后启动
            - 使用 tkinter.after 实现，线程安全
        """
        # 先取消之前的定时器（如果存在）
        self._stop_auto_refresh()
        
        # 启动新的定时器（5分钟后执行）
        self.auto_refresh_timer_id = self.root.after(
            self.auto_refresh_interval,
            self._auto_refresh_callback
        )
        
        print(f"[定时刷新] 已启动，将在 {self.auto_refresh_interval // 60000} 分钟后自动刷新")
    
    def _stop_auto_refresh(self):
        """
        停止定时自动刷新
        
        功能：
            - 取消当前的定时器
            - 在窗口关闭或重新启动定时器时调用
        """
        if self.auto_refresh_timer_id is not None:
            try:
                self.root.after_cancel(self.auto_refresh_timer_id)
                print("[定时刷新] 已停止")
            except:
                pass  # 忽略取消失败的情况
            finally:
                self.auto_refresh_timer_id = None
    
    def _auto_refresh_callback(self):
        """
        定时刷新回调函数
        
        功能：
            - 执行数据刷新
            - 刷新完成后重新启动定时器
        """
        print(f"[定时刷新] 开始自动刷新数据...")
        
        # 执行数据加载（会自动重启定时器）
        self.load_data()
    
    def run(self):
        """运行GUI主循环"""
        self.root.mainloop()
    
    def __del__(self):
        """
        析构函数，确保关闭所有数据库连接和停止定时器
        
        重要性:
            - 防止数据库连接泄漏
            - 释放系统资源
            - 避免文件锁定问题
            - 停止定时刷新任务
            
        注意:
            - 使用 try-except 防止析构时出错
            - 静默处理异常，避免影响程序退出
        """
        try:
            # 停止定时自动刷新
            self._stop_auto_refresh()
        except Exception:
            # 静默处理
            pass
        
        try:
            # 关闭 contact.db 连接
            if hasattr(self, 'contact_conn') and self.contact_conn:
                self.contact_conn.close()
        except Exception:
            # 静默处理，避免析构时抛出异常
            pass
        
        try:
            # 关闭主数据库连接（如果存在）
            if hasattr(self, 'conn') and self.conn:
                self.conn.close()
        except Exception:
            # 静默处理
            pass


def main():
    """主函数"""
    root = tk.Tk()
    app = DatabaseViewer(root)
    app.run()


if __name__ == "__main__":
    main()
