import warnings
from datetime import datetime
from pathlib import Path


# 下载目录：这里使用你指定的 Windows 下载目录。
DOWNLOAD_DIR = Path(r"C:\Users\LENOVO\Downloads")

# 需要读取的导出文件名前缀。
FILE_PREFIX = "收款单-批量导出收款订单-合并显示导出(合并一个文件)-"

# 控制预览行数，避免文件很大时一次性输出太多内容。
PREVIEW_ROWS = 20

# 用于排序的列名，必须和 Excel 表头保持一致。
TRADE_TIME_COLUMN = "交易时间"


def find_latest_export_file():
    """查找下载目录中以指定前缀开头、修改时间最新的 Excel 文件。"""
    matched_files = []
    excel_suffixes = [".xlsx", ".xlsm", ".xls"]

    for file_path in DOWNLOAD_DIR.iterdir():
        if (
            file_path.is_file()
            and file_path.name.startswith(FILE_PREFIX)
            and file_path.suffix.lower() in excel_suffixes
        ):
            matched_files.append(file_path)

    if not matched_files:
        return None

    return max(matched_files, key=lambda item: item.stat().st_mtime)


def is_not_empty(value):
    """判断单元格是否有真实内容，避免把 None 或空字符串当成有效数据。"""
    return value is not None and str(value).strip() != ""


def normalize_header(value):
    """统一表头文本格式，避免前后空格影响列名匹配。"""
    if value is None:
        return ""
    return str(value).strip()


def parse_trade_time(value):
    """把交易时间单元格转换成可排序的 datetime；无法解析时放到最后。"""
    if isinstance(value, datetime):
        return value

    if value is None:
        return datetime.min

    text = str(value).strip()
    if not text:
        return datetime.min

    time_formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]

    for time_format in time_formats:
        try:
            return datetime.strptime(text, time_format)
        except ValueError:
            continue

    return datetime.min


def get_non_empty_rows(worksheet):
    """读取工作表中的所有非空行，保留 Excel 原始行号方便排查。"""
    rows = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        row_values = list(row)

        if not any(is_not_empty(value) for value in row_values):
            continue

        rows.append((row_index, row_values))

    return rows


def print_sorted_rows_by_trade_time(worksheet):
    """按交易时间倒序输出当前工作表的数据。"""
    non_empty_rows = get_non_empty_rows(worksheet)

    if not non_empty_rows:
        print("这个工作表没有读取到非空数据。")
        return 0

    header_row_index, header = non_empty_rows[0]
    normalized_header = [normalize_header(value) for value in header]

    if TRADE_TIME_COLUMN not in normalized_header:
        print(f"未找到“{TRADE_TIME_COLUMN}”列，无法按交易时间排序。")
        print(f"当前识别到的表头：{normalized_header}")
        return 0

    trade_time_index = normalized_header.index(TRADE_TIME_COLUMN)
    data_rows = non_empty_rows[1:]

    if not data_rows:
        print("只读取到表头，没有读取到明细数据。")
        return 1

    sorted_rows = sorted(
        data_rows,
        key=lambda item: parse_trade_time(
            item[1][trade_time_index] if trade_time_index < len(item[1]) else None
        ),
        reverse=True,
    )

    print(f"表头位于第 {header_row_index} 行：{header}")
    print(f"按“{TRADE_TIME_COLUMN}”倒序后，预览前 {PREVIEW_ROWS} 条明细：")

    print_count = 0

    for original_row_index, row_values in sorted_rows[:PREVIEW_ROWS]:
        print(f"原第 {original_row_index} 行：{row_values}")
        print_count += 1

    return print_count + 1


def read_excel_file(file_path):
    """读取 Excel 文件，并打印所有工作表中的前几行非空数据。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("当前环境缺少 openpyxl，无法读取 Excel 文件。")
        print("请先执行：pip install openpyxl")
        return

    if file_path.suffix.lower() == ".xls":
        print("当前脚本使用 openpyxl 读取 Excel，暂不支持老式 .xls 文件。")
        print("请将文件另存为 .xlsx 后再运行脚本。")
        return

    # 部分系统导出的 Excel 缺少默认样式，openpyxl 会给出样式警告；
    # 这个警告不影响读取数据，这里屏蔽掉，避免干扰终端输出。
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style.*",
        category=UserWarning,
        module="openpyxl.styles.stylesheet",
    )

    # 该系统导出的 Excel 可能存在表格范围元数据不准确的问题。
    # 如果使用 read_only=True，openpyxl 可能只按错误范围读取到 A1，所以这里使用普通模式完整解析工作簿。
    workbook = load_workbook(file_path, read_only=False, data_only=True)

    print(f"工作表列表：{workbook.sheetnames}")

    total_printed_rows = 0

    for worksheet in workbook.worksheets:
        print("-" * 60)
        print(f"当前工作表：{worksheet.title}")
        print(f"工作表行数：{worksheet.max_row}，列数：{worksheet.max_column}")

        printed_rows = print_sorted_rows_by_trade_time(worksheet)
        total_printed_rows += printed_rows

    if total_printed_rows <= 1:
        print("-" * 60)
        print("只读取到表头，没有读取到明细数据。")
        print("如果你在 Excel 里能看到数据，通常是导出文件格式不标准，或数据不在普通单元格区域。")

    workbook.close()


def read_latest_export_file():
    """读取最新的收款单批量导出 Excel 文件。"""
    latest_file = find_latest_export_file()

    if latest_file is None:
        print(f"未找到 Excel 文件：{DOWNLOAD_DIR}\\{FILE_PREFIX}*.xlsx / *.xlsm / *.xls")
        return

    modified_time = datetime.fromtimestamp(latest_file.stat().st_mtime)

    print(f"找到最新文件：{latest_file}")
    print(f"文件修改时间：{modified_time:%Y-%m-%d %H:%M:%S}")
    print(f"开始预览前 {PREVIEW_ROWS} 行：")

    read_excel_file(latest_file)


def main():
    """脚本入口：读取下载目录中最新的收款单导出文件。"""
    read_latest_export_file()


if __name__ == "__main__":
    main()
