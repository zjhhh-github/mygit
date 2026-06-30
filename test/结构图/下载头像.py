import openpyxl
import requests
from io import BytesIO
from PIL import Image
import os

def download_images_from_excel(excel_path, output_folder):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"正在读取Excel文件: {excel_path}")
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"创建输出文件夹: {output_folder}")
    
    success_count = 0
    fail_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=5, min_row=1), start=1):
        c_cell = ws[f'C{row_idx}']
        e_cell = ws[f'E{row_idx}']
        
        c_value = c_cell.value
        e_value = e_cell.value
        
        if c_value and e_value:
            filename = str(c_value).strip()
            url = str(e_value).strip()
            
            if url.startswith(('http://', 'https://')):
                # print(f"\n处理第 {row_idx} 行:")
                # print(f"  文件名: {filename}")
                # print(f"  URL: {url}")
                
                try:
                    response = requests.get(url, timeout=10)
                    if response.status_code == 200:
                        img_data = BytesIO(response.content)
                        img = Image.open(img_data)
                        
                        save_path = os.path.join(output_folder, f"{filename}.png")
                        img.save(save_path)
                        
                        # print(f"  成功保存到: {save_path}")
                        success_count += 1
                    else:
                        print(f"  下载失败，状态码: {response.status_code}")
                        fail_count += 1
                except Exception as e:
                    print(f"  处理失败: {e}")
                    fail_count += 1
            else:
                print(f"第 {row_idx} 行E列不是有效的URL: {url}")
                fail_count += 1
    
    print(f"\n完成！成功下载 {success_count} 张图片，失败 {fail_count} 张")
    print(f"图片保存在: {output_folder}")

if __name__ == "__main__":
    excel_path = r"C:\Users\LENOVO\Desktop\通讯录导出.xlsx"
    output_folder = r"C:\Users\LENOVO\Desktop\123"
    
    download_images_from_excel(excel_path, output_folder)
