import openpyxl
import requests
from io import BytesIO
from PIL import Image
import os
from openpyxl.drawing.image import Image as ExcelImage

def insert_images_from_urls(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print("正在读取E列的图片URL...")
    
    temp_files = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_col=5, max_col=5, min_row=1), start=1):
        cell = row[0]
        if cell.value:
            url = str(cell.value).strip()
            
            if url.startswith(('http://', 'https://')):
                print(f"\n处理第 {row_idx} 行: {url}")
                
                try:
                    response = requests.get(url, timeout=10)
                    if response.status_code == 200:
                        img_data = BytesIO(response.content)
                        img = Image.open(img_data)
                        
                        img_path = f"temp_image_{row_idx}.png"
                        img.save(img_path)
                        temp_files.append(img_path)
                        
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 100
                        excel_img.height = 100
                        
                        ws.add_image(excel_img, f'G{row_idx}')
                        
                        ws[f'G{row_idx}'].value = url
                        
                        print(f"成功插入图片到 G{row_idx}")
                    else:
                        print(f"下载失败，状态码: {response.status_code}")
                        ws[f'G{row_idx}'].value = f"下载失败: {response.status_code}"
                except Exception as e:
                    print(f"处理失败: {e}")
                    ws[f'G{row_idx}'].value = f"错误: {str(e)}"
            else:
                print(f"第 {row_idx} 行不是有效的URL: {url}")
    
    output_path = excel_path.replace('.xlsx', '_with_images.xlsx')
    wb.save(output_path)
    print(f"\n文件已保存到: {output_path}")
    
    print("\n正在清理临时文件...")
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
            print(f"已删除: {temp_file}")
        except Exception as e:
            print(f"删除失败 {temp_file}: {e}")
    
    return output_path

if __name__ == "__main__":
    excel_path = r"C:\Users\LENOVO\Desktop\通讯录导出.xlsx"
    insert_images_from_urls(excel_path)
