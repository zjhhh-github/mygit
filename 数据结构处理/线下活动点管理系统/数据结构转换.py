import json
from pathlib import Path
import re
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook


# 优先使用你当前指定的输入文件；若不存在则自动回退到脚本目录内最近的 xlsx 文件
默认输入路径 = Path(r"C:\Users\LENOVO\Desktop\线下活动点.xlsx")
默认工作表名 = "线下活动点"
默认输出路径 = Path(__file__).resolve().parent / "线下活动点管理系统上传数据.json"
# 按名称匹配回填：从下载 JSON 回填到 Excel K 列（第 11 列）
回填JSON路径 = Path(r"C:\Users\LENOVO\Downloads\活动点数据_2026-04-09.json")
启用回填K列 = True

# ==================== JSON 结构自定义配置区 ====================
# 说明：
# 1) 将模板中的 "{{列名}}" 写成 Excel 表头名，即可把该列值映射到对应 JSON 字段；
# 2) 不是 "{{列名}}" 的普通字符串会按固定值输出；
# 3) 支持嵌套字典和列表，便于自定义复杂 JSON 结构。
启用自定义结构 = True
JSON结构模板 = {
    "名称": "{{名称}}",
    "省": "{{省}}",
    "市": "{{市}}",
    "区": "{{区县}}",
    "详细地址": "{{详细地址}}",
    # 注意这里必须是字符串或字典，不能写成 set（否则无法 JSON 序列化）
    "经纬度": "{{经纬度}}",
    # 绑定教务账号支持多列兜底：按从左到右的优先级取首个非空值
    "绑定教务账号": "{{绑定教务账号|绑定教务|Unnamed:9|Unnamed: 9}}"
}


def 定位输入文件() -> Path:
    """
    定位要转换的 Excel 文件。
    规则：
    1) 优先使用默认输入路径；
    2) 若不存在，则从脚本同目录选择最新修改的 xlsx 文件。
    """
    if 默认输入路径.exists():
        return 默认输入路径

    当前目录 = Path(__file__).resolve().parent
    候选文件 = sorted(当前目录.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if 候选文件:
        return 候选文件[0]

    raise FileNotFoundError("未找到可用的 Excel 文件，请检查默认路径或将 xlsx 放到脚本同目录。")


def 应用模板值(模板值, 行数据: dict):
    """
    递归应用 JSON 模板，生成单行目标结构。

    规则：
    - "{{列名}}"：从当前行读取对应列值，列不存在时返回空字符串
    - dict/list：递归处理
    - 其他类型：作为固定值原样输出
    """
    def 标准化字段值(原始值):
        """
        统一字段输出格式：
        - 空值输出空字符串
        - 浮点整数去掉“.0”（如 3676.0 -> 3676）
        - 其余类型转字符串并去前后空白
        """
        if pd.isna(原始值):
            return ""
        文本值 = str(原始值).strip()
        if 文本值.lower() == "nan":
            return ""
        if 文本值.endswith(".0") and 文本值[:-2].isdigit():
            return 文本值[:-2]
        return 文本值

    def 解析绑定教务账号(文本值):
        """
        解析“绑定教务账号”字段，支持一个或多个账号。
        - 多账号分隔符：中文/英文逗号、顿号、斜杠、分号、空白
        - 纯数字统一补齐为 6 位（111 -> 000111）
        返回：
        - 始终返回字符串列表（单账号也返回单元素列表）
        """
        分段列表 = [片段.strip() for 片段 in re.split(r"[、,，/;；\s]+", 文本值) if 片段.strip()]
        if not 分段列表:
            return ""

        账号列表 = []
        for 片段 in 分段列表:
            if 片段.endswith(".0") and 片段[:-2].isdigit():
                片段 = 片段[:-2]
            if 片段.isdigit():
                片段 = 片段.zfill(6)
            账号列表.append(片段)

        # 去重并保持原顺序，避免重复账号污染结果
        去重账号列表 = list(dict.fromkeys(账号列表))
        return 去重账号列表

    if isinstance(模板值, str):
        if 模板值.startswith("{{") and 模板值.endswith("}}"):
            占位内容 = 模板值[2:-2].strip()
            # 支持“多列兜底”：{{列A|列B|列C}}
            候选列名列表 = [列名.strip() for 列名 in 占位内容.split("|") if 列名.strip()]
            for 列名 in 候选列名列表:
                字段值 = 标准化字段值(行数据.get(列名, ""))
                if 字段值:
                    # 业务规则：绑定教务字段支持多账号，并统一补齐为 6 位账号
                    if 列名 in {"绑定教务账号", "绑定教务", "Unnamed:9", "Unnamed: 9"}:
                        # 这里不能提前 return，需要把多个候选列都汇总，避免丢失第二个教务账号
                        汇总账号列表 = []
                        for 绑定列名 in 候选列名列表:
                            绑定列值 = 标准化字段值(行数据.get(绑定列名, ""))
                            if not 绑定列值:
                                continue
                            解析结果 = 解析绑定教务账号(绑定列值)
                            if isinstance(解析结果, list):
                                汇总账号列表.extend(解析结果)
                            elif 解析结果:
                                汇总账号列表.append(解析结果)
                        # 去重并保持顺序
                        return list(dict.fromkeys(汇总账号列表))
                    return 字段值
            return ""
        return 模板值
    if isinstance(模板值, dict):
        return {键: 应用模板值(值, 行数据) for 键, 值 in 模板值.items()}
    if isinstance(模板值, list):
        return [应用模板值(项, 行数据) for 项 in 模板值]
    return 模板值


def 生成输出记录列表(数据表: pd.DataFrame) -> List[Dict]:
    """
    根据配置生成输出记录列表。
    - 启用自定义结构时：按 JSON结构模板 输出
    - 关闭时：按原始表头直接输出
    """
    行记录列表 = 数据表.to_dict(orient="records")
    if not 启用自定义结构:
        return 行记录列表

    输出记录列表 = []
    for 行数据 in 行记录列表:
        输出记录列表.append(应用模板值(JSON结构模板, 行数据))
    return 输出记录列表


def 回填JSON到Excel最后一列(json_path: Path, excel_path: Path, 工作表名: str) -> None:
    """
    读取 JSON 后按“名称”匹配，回填到 Excel K 列（第 11 列）。

    规则：
    - JSON 中读取字段：名称、经纬度
    - 仅覆盖命中名称的行，未命中保持原值不动
    """
    if not json_path.exists():
        raise FileNotFoundError(f"未找到 JSON 文件：{json_path}")
    if not excel_path.exists():
        raise FileNotFoundError(f"未找到 Excel 文件：{excel_path}")

    with json_path.open("r", encoding="utf-8") as f:
        json数据 = json.load(f)

    if not isinstance(json数据, list):
        raise ValueError("JSON 顶层结构必须是数组(list)")

    def 格式化经纬度(坐标文本: str) -> str:
        """将 '经度,纬度' 格式的坐标文本中经纬度各保留 6 位小数；格式异常时原样返回。"""
        部分列表 = 坐标文本.split(",")
        if len(部分列表) != 2:
            return 坐标文本
        try:
            经度 = round(float(部分列表[0].strip()), 6)
            纬度 = round(float(部分列表[1].strip()), 6)
            return f"{经度:.6f},{纬度:.6f}"
        except ValueError:
            return 坐标文本

    名称映射坐标文本 = {}
    for 项 in json数据:
        if not isinstance(项, dict):
            continue
        名称 = str(项.get("名称", "")).strip()
        坐标值 = str(项.get("经纬度", "")).strip()
        if not 名称:
            continue
        # 经纬度只保留小数点后 6 位
        名称映射坐标文本[名称] = 格式化经纬度(坐标值)

    工作簿 = load_workbook(str(excel_path))
    if 工作表名 not in 工作簿.sheetnames:
        raise ValueError(f"Excel 中未找到工作表：{工作表名}")
    工作表 = 工作簿[工作表名]

    # 按第一行表头定位“名称”列
    名称列索引 = None
    for 列号 in range(1, 工作表.max_column + 1):
        表头 = str(工作表.cell(row=1, column=列号).value or "").strip()
        if 表头 == "名称":
            名称列索引 = 列号
            break
    if 名称列索引 is None:
        raise ValueError("Excel 表头中未找到“名称”列，无法按名称匹配回填")

    目标列索引 = 11  # K 列
    匹配成功数 = 0
    未匹配名称数 = 0
    for 行号 in range(2, 工作表.max_row + 1):
        名称值 = str(工作表.cell(row=行号, column=名称列索引).value or "").strip()
        if not 名称值:
            continue
        if 名称值 in 名称映射坐标文本:
            工作表.cell(row=行号, column=目标列索引).value = 名称映射坐标文本[名称值]
            匹配成功数 += 1
        else:
            未匹配名称数 += 1

    # 若原文件被占用，则自动回退到时间戳文件，避免中断
    try:
        工作簿.save(str(excel_path))
        实际输出路径 = excel_path
    except PermissionError:
        回退路径 = excel_path.with_name(
            f"{excel_path.stem}_回填_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
        )
        工作簿.save(str(回退路径))
        实际输出路径 = 回退路径

    print(f"回填完成：{实际输出路径}")
    print(f"回填列：K列(第{目标列索引}列)")
    print(f"按名称匹配成功：{匹配成功数} 行")
    print(f"未匹配名称：{未匹配名称数} 行")


def main() -> None:
    """
    将 Excel 表格内容转换为 JSON（按行转字典）。
    """
    输入路径 = 定位输入文件()
    # 按配置读取指定工作表；空值统一为 "" 便于下游消费
    数据表 = pd.read_excel(输入路径, sheet_name=默认工作表名).fillna("")
    记录列表 = 生成输出记录列表(数据表)

    with 默认输出路径.open("w", encoding="utf-8") as f:
        json.dump(记录列表, f, ensure_ascii=False, indent=2)

    print(f"转换完成：{输入路径}")
    print(f"输出文件：{默认输出路径}")
    print(f"总行数：{len(记录列表)}")

    # 按你的需求：读取指定 JSON，按“名称”匹配回填到 Excel K 列
    if 启用回填K列:
        回填JSON到Excel最后一列(回填JSON路径, 输入路径, 默认工作表名)


if __name__ == "__main__":
    main()