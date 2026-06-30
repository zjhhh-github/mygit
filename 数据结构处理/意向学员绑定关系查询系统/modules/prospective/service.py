# -*- coding: utf-8 -*-
"""
意向查询系统 —— 服务层（Phase 2 雏形）
=======================================

本文件是逐步成形的「业务服务层」入口。

当前阶段（Phase 2b）只承载一项能力：
    - export_to_excel(data, output_dir, timestamp)
      把 12 列业务数据写成「意向专用通讯录_YYYYMMDD_HHMMSS.xlsx」文件。
      行为完全复刻 db_viewer.DatabaseViewer._do_export_full（手动导出/自动导出共用）。

后续阶段会再补：
    - load_all(config) -> LoadResult
      把 _load_data_in_thread 的业务部分搬过来，返回结构化数据 + 降级消息

设计要点：
    - 函数无 Tk 依赖，可在工作线程中调用
    - 异常不抛出，统一以 (ok, path_or_name, error) 三元组返回
    - 列索引常量从 modules.prospective.config 读取
"""

import gc
import os

from modules.prospective.config import COL_OBJ_IS_DELETE


# Excel 导出表头（"对象"统一改为"意向学员"）
EXPORT_HEADERS = [
    "意向学员(昵称)", "意向学员(微信ID)", "意向学员(微信号)", "意向学员(总微信号)",
    "意向学员(添加时间)", "意向学员(内部备注)", "意向学员(是否删除)",
    "来源(昵称)", "来源(微信ID)", "来源(微信号)", "来源(总微信号)", "来源(内部备注)",
]

# 文件名前缀（不含时间戳与扩展名）
EXPORT_FILENAME_PREFIX = "意向专用通讯录导出"


def export_to_excel(data, timestamp, output_dir=None):
    """
    把 12 列业务数据导出为 Excel 文件。

    Args:
        data: list[dict]，每项结构 {'values': tuple|list[12], 'tag': str, ...}
        timestamp: 时间戳字符串（建议格式 'YYYYMMDD_HHMMSS' 或带毫秒）
        output_dir: 可选输出目录绝对路径
            - None：保存到当前工作目录，第二个返回值为文件名
            - 非空：保存到该目录下，调用方需保证目录已存在且可写，
                    第二个返回值为完整路径

    Returns:
        (ok, path_or_name, error)
            ok            : bool
            path_or_name  : 文件名（output_dir=None）或完整路径（指定 output_dir）
            error         : 失败原因；成功为空字符串

    实现要点：
        - 延迟导入 openpyxl，避免主程序启动时为此付费
        - "是否删除"列做状态转换（已删除→✅，否则→❌），其余列原值导出
        - 表头微软雅黑加粗、灰色背景、冻结表头
        - 列宽 / 行高统一 19.5
        - 任何异常均通过返回值传递，不抛出
    """
    # 延迟导入 openpyxl（仅在导出时导入，提升启动速度）
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    filename = "{}_{}.xlsx".format(EXPORT_FILENAME_PREFIX, timestamp)

    try:
        # 安全性检查：确保文件名不包含路径分隔符（防止路径遍历）
        if '/' in filename or '\\' in filename or '..' in filename:
            return False, filename, "文件名包含非法字符！"

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "意向专用通讯录"

        # 写入表头
        ws.append(EXPORT_HEADERS)

        # 冻结表头（第一行）
        ws.freeze_panes = 'A2'

        # 表头样式（微软雅黑、11 号、加粗、居中、黑字、灰色-25% 背景）
        header_font = Font(name="微软雅黑", bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        for col_idx, _ in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # 写入数据
        for item in data:
            values = item['values']
            cleaned_values = [
                ("✅" if str(val) == "已删除" else "❌")
                if idx == COL_OBJ_IS_DELETE
                else str(val) if val is not None else ""
                for idx, val in enumerate(values)
            ]
            ws.append(cleaned_values)

        # 列宽统一 19.5
        for col_idx in range(1, len(EXPORT_HEADERS) + 1):
            col_letter = chr(64 + col_idx)  # A, B, C, ...
            ws.column_dimensions[col_letter].width = 19.5

        # 行高统一 19.5（包括表头与数据行）
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 19.5

        # 数据行样式（批量 iter_rows 更快）
        data_font = Font(name="微软雅黑", color="000000", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=len(EXPORT_HEADERS)):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment

        # 保存文件：默认当前目录；指定 output_dir 则拼接保存
        save_path = os.path.join(output_dir, filename) if output_dir else filename
        wb.save(save_path)

        # 释放工作簿对象，减少内存占用
        del wb
        del ws
        gc.collect()

        return True, save_path, ""

    except PermissionError:
        return False, filename, "文件 {} 正在被其他程序占用！请关闭该文件后重试。".format(filename)
    except Exception as e:
        return False, filename, "{}: {}".format(type(e).__name__, e)
