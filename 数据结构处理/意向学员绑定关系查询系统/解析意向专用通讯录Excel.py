import json
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# 源文件：意向专用通讯录导出的 Excel。
INPUT_EXCEL = Path(r"C:\Users\LENOVO\Downloads\意向专用通讯录导出_20260519_115240_255.xlsx")

# 输出文件：结构与桌面已有的 _脚本输出_1_解析结果.json 保持一致。
OUTPUT_JSON = Path(r"C:\Users\LENOVO\Desktop\意向专用通讯录导出_解析结果.json")


def clean_value(value):
    """把 Excel 单元格值统一转成去掉首尾空格的字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def format_date_to_yyyymmdd(value):
    """
    将 Excel 中的添加时间转换成 yyyyMMdd。

    Excel 样例中时间格式类似：2025/08/29 14:33:05。
    如果后续出现 datetime 类型，也会按同样规则处理。
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")

    text = clean_value(value)
    if not text:
        return ""

    for date_format in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).strftime("%Y%m%d")
        except ValueError:
            continue

    # 日期格式无法识别时，保留空值，避免写入错误日期。
    return ""


def build_header_index(header_row):
    """根据表头名称建立列索引，避免代码依赖固定列号。"""
    header_index = {}

    for index, value in enumerate(header_row):
        header_name = clean_value(value)
        if header_name:
            header_index[header_name] = index

    return header_index


def get_cell(row, header_index, header_name):
    """按表头名称安全读取当前行的单元格值。"""
    index = header_index.get(header_name)
    if index is None or index >= len(row):
        return ""
    return clean_value(row[index])


def parse_excel(input_excel: Path):
    """
    解析意向专用通讯录 Excel，输出目标嵌套结构。

    字段映射：
    - 意向学员(总微信号) -> 意向学员微信号
    - 意向学员(微信ID) -> 意向学员微信原始ID
    - 来源(总微信号) -> 推荐人总微信号
    - 来源(微信ID) -> 推荐人微信原始ID
    - 意向学员(添加时间) -> 绑定日期

    Excel 中没有“是否报名 / 解绑日期 / 绑定状态”字段：
    - 是否报名：默认写“未报名”
    - 解绑日期：默认写空字符串
    - 绑定状态：有来源时写“有绑定”
    """
    if not input_excel.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{input_excel}")

    workbook = load_workbook(str(input_excel), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    # 不传 row_offset / values_only，兼容不同版本的 openpyxl。
    rows = sheet.iter_rows()
    header_row = [cell.value for cell in next(rows)]
    header_index = build_header_index(header_row)

    students = OrderedDict()
    row_count = 0
    recommend_count = 0
    skipped_empty_student = 0

    for excel_row in rows:
        row_count += 1
        row = [cell.value for cell in excel_row]

        student_wechat = get_cell(row, header_index, "意向学员(总微信号)")
        student_original_id = get_cell(row, header_index, "意向学员(微信ID)")
        add_time = row[header_index["意向学员(添加时间)"]] if "意向学员(添加时间)" in header_index else None
        bind_date = format_date_to_yyyymmdd(add_time)

        referrer_wechat = get_cell(row, header_index, "来源(总微信号)")
        referrer_original_id = get_cell(row, header_index, "来源(微信ID)")

        if not student_wechat:
            skipped_empty_student += 1
            continue

        if student_wechat not in students:
            students[student_wechat] = {
                "意向学员微信号": student_wechat,
                "意向学员微信原始ID": student_original_id,
                "是否报名": "未报名",
                "推荐": [],
            }
        else:
            # 同一个学员重复出现时，优先补齐缺失的原始 ID。
            if not students[student_wechat]["意向学员微信原始ID"] and student_original_id:
                students[student_wechat]["意向学员微信原始ID"] = student_original_id

        if referrer_wechat:
            students[student_wechat]["推荐"].append(
                {
                    "推荐人总微信号": referrer_wechat,
                    "推荐人微信原始ID": referrer_original_id,
                    "绑定日期": bind_date,
                    "绑定周期": None,
                    "解绑日期": "",
                    "绑定状态": "有绑定",
                }
            )
            recommend_count += 1

    workbook.close()

    stats = {
        "读取 Excel 数据行数": row_count,
        "输出意向学员数量": len(students),
        "输出推荐关系数量": recommend_count,
        "跳过空学员微信号行数": skipped_empty_student,
    }

    return list(students.values()), stats


def main():
    """脚本入口：解析 Excel 并写入 JSON 文件。"""
    result, stats = parse_excel(INPUT_EXCEL)

    with OUTPUT_JSON.open("w", encoding="utf-8") as file:
        json.dump(result, file, ensure_ascii=False, indent=2)

    print("解析完成")
    for name, value in stats.items():
        print(f"{name}：{value}")
    print(f"输出文件：{OUTPUT_JSON}")


if __name__ == "__main__":
    main()
