"""
校区数据汇总脚本（月度 + 周度）

流失学员：上一期有、当前期没有的学员（课程卡快照对比）
新增学员：当前期有、上一期没有的学员（课程卡快照对比）
月报和周报各自独立对比，不互相干扰。
"""

import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
from collections import defaultdict


# ─────────────────────────────────────────────
# 校区编号映射表
# ─────────────────────────────────────────────
CAMPUS_NUMBER_MAP = {
    "0001": "万悦城",
    "0002": "万锦合泰",
    "0003": "麦迪逊",
    "0004": "恒大城",
    "0005": "购觅",
    "0006": "文苑",
    "0007": "思德",
    "0008": "银泰",
    "0009": "西万达",
    "0010": "和林",
    "0011": "秀水",
    "0012": "鼓楼",
    "0013": "万锦观悦",
    "0014": "学院里",
    "0015": "苏宁",
    "0016": "昆北",
    "0017": "中海金地",
    "0018": "东河湾",
    "0019": "长安金座",
    "0020": "太伟方恒",
    "0021": "帅家营",
    "0022": "毫沁营",
    "0023": "五里营",
    "0024": "盛乐",
    "0025": "东胜未来世界",
    "0026": "新雅",
    "0027": "托县",
    "0028": "金桥",
    "0029": "范家营",
    "0030": "卓育",
    "0031": "包百",
}


def get_campus_number(campus_name: str) -> str:
    for number, keyword in CAMPUS_NUMBER_MAP.items():
        if keyword in campus_name:
            return number
    return ""


# ─────────────────────────────────────────────
# 配置
# ─────────────────────────────────────────────
FOLDERS = {
    "包头": r"C:\Users\LENOVO\Desktop\包头",
    "呼和浩特+鄂尔多斯": r"C:\Users\LENOVO\Desktop\呼和浩特+鄂尔多斯",
}
OUTPUT_PATH = r"C:\Users\LENOVO\Desktop\校区月度汇总报表.xlsx"


# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────
def get_date_range(xl):
    df = xl.parse("统计日期区间说明", header=None)
    desc = str(df.iloc[0, 0])
    m = re.search(r"(\d{4}-\d{2}-\d{2})至(\d{4}-\d{2}-\d{2})", desc)
    if m:
        start = datetime.strptime(m.group(1), "%Y-%m-%d")
        end   = datetime.strptime(m.group(2), "%Y-%m-%d")
        return start, end
    return None, None


def is_monthly(xl):
    s, e = get_date_range(xl)
    return bool(s and e and (e - s).days >= 27)


def period_label(start, end, monthly):
    if monthly:
        return start.strftime("%Y-%m")
    return f"{start.strftime('%Y-%m-%d')}~{end.strftime('%m-%d')}"


# ─────────────────────────────────────────────
# 读取课程卡快照：{校区: set(学员姓名)}，仅状态="正常"
# ─────────────────────────────────────────────
def get_snapshot(xl):
    result = defaultdict(set)
    if "课程卡" not in xl.sheet_names:
        return dict(result)
    df = xl.parse("课程卡")
    if not {"课程卡状态", "所在校区", "学员姓名"}.issubset(df.columns):
        return dict(result)
    active = df[df["课程卡状态"] == "正常"]
    for _, row in active.iterrows():
        campus = row["所在校区"]
        name   = row["学员姓名"]
        if isinstance(campus, str) and isinstance(name, str):
            result[campus].add(name)
    return dict(result)


# ─────────────────────────────────────────────
# 提取单个文件的指标（不含新增/流失，由主流程填入）
# ─────────────────────────────────────────────
def extract_metrics(xl, fee_prefix):
    sheets = xl.sheet_names
    campus_data = defaultdict(lambda: {
        "试听人数":    0,
        "退费订单":    0,
        "老带新营收":  0,
        "新增学员":    0,   # 由主流程填入
        "在读学员":    0,
        "流失学员":    0,   # 由主流程填入
        "流失学员姓名": "",
        "转线上学员":  0,
        "课消总课时":  0,
        "请假课时":    0,
        "剩余课时":    0,
        "冻结课时":    0,
        "平均出勤率":  None,
    })

    class_sheet  = f"{fee_prefix}上课记录"
    fee_sheet    = f"{fee_prefix}缴费记录"
    attend_sheet = f"{fee_prefix}出勤记录"

    # 1. 试听人数
    if class_sheet in sheets:
        df_class = xl.parse(class_sheet)
        if {"上课学员", "所在校区"}.issubset(df_class.columns):
            for _, row in df_class.iterrows():
                campus = row.get("所在校区")
                info   = str(row.get("上课学员", ""))
                if not isinstance(campus, str):
                    continue
                trial = set()
                for part in info.split("\n"):
                    part = part.strip().rstrip("，,")
                    if "|试听" in part:
                        name = part.split("|")[0].strip()
                        if name:
                            trial.add(name)
                campus_data[campus]["_试听集"] = (
                    campus_data[campus].get("_试听集", set()) | trial
                )
    for campus in campus_data:
        if "_试听集" in campus_data[campus]:
            campus_data[campus]["试听人数"] = len(campus_data[campus]["_试听集"])

    # 2. 缴费记录
    if fee_sheet in sheets:
        df_fee = xl.parse(fee_sheet)
        for _, row in df_fee.iterrows():
            campus = row.get("所在校区")
            if not isinstance(campus, str):
                continue
            fee_type   = str(row.get("缴费类型", ""))
            fee_amount = float(row.get("缴费金额", 0) or 0)
            pkg_name   = str(row.get("套餐名称", ""))
            remark     = str(row.get("备注", ""))
            if fee_type == "退费":
                campus_data[campus]["退费订单"] += 1
            if fee_type in ("报名", "续报") and any(
                kw in remark for kw in ["老带新", "转介绍", "推荐"]
            ):
                campus_data[campus]["老带新营收"] += fee_amount
            if "线上" in pkg_name and fee_type in ("报名", "续报"):
                campus_data[campus]["转线上学员"] += 1

    # 3. 在读学员 / 剩余课时 / 冻结课时
    if "课程卡" in sheets:
        df_card = xl.parse("课程卡")
        if {"课程卡状态", "所在校区", "学员姓名", "剩余课时"}.issubset(df_card.columns):
            active = df_card[df_card["课程卡状态"] == "正常"]
            frozen = df_card[df_card["课程卡状态"] == "停用"]
            for campus, grp in active.groupby("所在校区"):
                if isinstance(campus, str):
                    campus_data[campus]["在读学员"] = grp["学员姓名"].nunique()
                    campus_data[campus]["剩余课时"] = grp["剩余课时"].sum()
            for campus, grp in frozen.groupby("所在校区"):
                if isinstance(campus, str):
                    campus_data[campus]["冻结课时"] = grp["剩余课时"].sum()

    # 4. 课消总课时 / 平均出勤率
    if class_sheet in sheets:
        df_class = xl.parse(class_sheet)
        if {"实扣课时", "所在校区"}.issubset(df_class.columns):
            for campus, grp in df_class.groupby("所在校区"):
                if isinstance(campus, str):
                    campus_data[campus]["课消总课时"] += grp["实扣课时"].sum()
        if {"到课人数", "请假人数", "旷课人数", "所在校区"}.issubset(df_class.columns):
            for campus, grp in df_class.groupby("所在校区"):
                if isinstance(campus, str):
                    total = (grp["到课人数"].sum()
                             + grp["请假人数"].sum()
                             + grp["旷课人数"].sum())
                    if total > 0:
                        campus_data[campus]["平均出勤率"] = round(
                            grp["到课人数"].sum() / total * 100, 2
                        )

    # 5. 请假课时（请假行数即请假课时，因请假不扣课）
    if attend_sheet in sheets:
        df_att = xl.parse(attend_sheet)
        if {"出勤状态", "所在校区"}.issubset(df_att.columns):
            leave = df_att[df_att["出勤状态"] == "请假"]
            for campus, grp in leave.groupby("所在校区"):
                if isinstance(campus, str):
                    campus_data[campus]["请假课时"] += len(grp)

    # 清理临时字段
    for campus in campus_data:
        campus_data[campus].pop("_试听集", None)

    return dict(campus_data)


# ─────────────────────────────────────────────
# 根据相邻快照填充新增 / 流失
# 流失 = 上一期有、当前期没有
# 新增 = 当前期有、上一期没有
# ─────────────────────────────────────────────
def fill_new_and_lost(i, snapshots, campus_metrics):
    curr_snap = snapshots[i]
    prev_snap = snapshots[i - 1] if i > 0 else {}

    for campus, metrics in campus_metrics.items():
        curr_stu = curr_snap.get(campus, set())
        prev_stu = prev_snap.get(campus, set())

        lost  = prev_stu - curr_stu
        added = curr_stu - prev_stu

        metrics["流失学员"]     = len(lost)
        metrics["流失学员姓名"] = "、".join(sorted(lost)) if lost else ""
        metrics["新增学员"]     = len(added)


# ─────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────
def main():
    monthly_records = []
    weekly_records  = []

    for city_name, folder_path in FOLDERS.items():
        print(f"\n处理文件夹：{city_name}")

        # 收集所有文件信息，按开始日期排序
        file_infos = []
        for fname in sorted(os.listdir(folder_path)):
            if fname.startswith("~") or not fname.endswith(".xlsx"):
                continue
            fpath = os.path.join(folder_path, fname)
            try:
                xl = pd.ExcelFile(fpath)
            except Exception as e:
                print(f"  跳过：{fname} - {e}")
                continue
            s, e = get_date_range(xl)
            if s is None:
                continue
            file_infos.append((s, e, fpath, is_monthly(xl), fname))
        file_infos.sort(key=lambda x: x[0])

        # 月报和周报分别建立列表
        monthly_infos = [(s, e, fp, fn) for s, e, fp, im, fn in file_infos if im]
        weekly_infos  = [(s, e, fp, fn) for s, e, fp, im, fn in file_infos if not im]

        # 读取快照序列（月报一套，周报一套）
        monthly_snapshots = [get_snapshot(pd.ExcelFile(fp)) for s, e, fp, fn in monthly_infos]
        weekly_snapshots  = [get_snapshot(pd.ExcelFile(fp)) for s, e, fp, fn in weekly_infos]

        # 处理月报
        for i, (s, e, fpath, fname) in enumerate(monthly_infos):
            print(f"  [月报] {fname}  {s.date()} ~ {e.date()}")
            xl = pd.ExcelFile(fpath)
            campus_metrics = extract_metrics(xl, "月")
            fill_new_and_lost(i, monthly_snapshots, campus_metrics)
            label = period_label(s, e, True)
            for campus, metrics in campus_metrics.items():
                record = {
                    "编号": get_campus_number(campus),
                    "城市": city_name,
                    "校区": campus,
                    "期次": label,
                }
                record.update(metrics)
                monthly_records.append(record)

        # 处理周报
        for i, (s, e, fpath, fname) in enumerate(weekly_infos):
            print(f"  [周报] {fname}  {s.date()} ~ {e.date()}")
            xl = pd.ExcelFile(fpath)
            campus_metrics = extract_metrics(xl, "周")
            fill_new_and_lost(i, weekly_snapshots, campus_metrics)
            label = period_label(s, e, False)
            for campus, metrics in campus_metrics.items():
                record = {
                    "编号": get_campus_number(campus),
                    "城市": city_name,
                    "校区": campus,
                    "期次": label,
                }
                record.update(metrics)
                weekly_records.append(record)

    # ─────────────────────────────────────────
    # 整理 DataFrame
    # ─────────────────────────────────────────
    indicator_cols = [
        "试听人数",
        "退费订单",
        "老带新营收",
        "新增学员",
        "在读学员",
        "流失学员",
        "流失学员姓名",
        "转线上学员",
        "课消总课时",
        "请假课时",
        "剩余课时",
        "冻结课时",
        "平均出勤率",
    ]

    def build_df(records):
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        for col in indicator_cols:
            if col not in df.columns:
                df[col] = 0
        df = df[["编号", "城市", "校区", "期次"] + indicator_cols].copy()
        int_cols = [c for c in indicator_cols
                    if c not in ("平均出勤率", "流失学员姓名", "老带新营收")]
        for col in int_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
        df["平均出勤率"] = pd.to_numeric(df["平均出勤率"], errors="coerce").round(2)
        df["老带新营收"] = pd.to_numeric(df["老带新营收"], errors="coerce").fillna(0).round(2)
        df["_sort"] = df["编号"].apply(lambda x: int(x) if x.isdigit() else 9999)
        df = (df.sort_values(["_sort", "期次"])
                .drop(columns=["_sort"])
                .reset_index(drop=True))
        df["编号"] = df["编号"].astype(str)
        return df

    df_monthly = build_df(monthly_records)
    df_weekly  = build_df(weekly_records)

    print(f"\n月度记录：{len(df_monthly)} 条")
    print(f"周度记录：{len(df_weekly)} 条")

    # ─────────────────────────────────────────
    # 写入 Excel
    # ─────────────────────────────────────────
    print(f"\n正在写入：{OUTPUT_PATH}")

    def write_sheet(wb, sheet_name, df):
        if df is None or df.empty:
            return
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(list(df.columns))
        for row_vals in df.itertuples(index=False, name=None):
            ws.append(list(row_vals))
        if "编号" in df.columns:
            col_idx = list(df.columns).index("编号") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "@"

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    write_sheet(wb, "月度总览", df_monthly)
    write_sheet(wb, "周度总览", df_weekly)

    for city in df_monthly["城市"].unique():
        write_sheet(wb, f"{city}（月）"[:31], df_monthly[df_monthly["城市"] == city])
    for city in df_weekly["城市"].unique():
        write_sheet(wb, f"{city}（周）"[:31], df_weekly[df_weekly["城市"] == city])

    all_campuses = list(df_monthly["校区"].unique()) + [
        c for c in df_weekly["校区"].unique()
        if c not in df_monthly["校区"].values
    ]
    for campus in all_campuses:
        df_m = df_monthly[df_monthly["校区"] == campus] if not df_monthly.empty else pd.DataFrame()
        if not df_m.empty:
            write_sheet(wb, campus[:31], df_m)

    wb.save(OUTPUT_PATH)
    print(f"完成！输出：{OUTPUT_PATH}")

    # 验证预览
    print("\n月度数据预览（万悦城）：")
    if not df_monthly.empty:
        wanyu = df_monthly[df_monthly["校区"] == "万悦城"][
            ["期次", "在读学员", "新增学员", "流失学员"]
        ]
        print(wanyu.to_string(index=False))


if __name__ == "__main__":
    main()
