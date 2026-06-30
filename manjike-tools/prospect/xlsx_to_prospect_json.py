#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""意向通讯录 xlsx → 标准 JSON。

【脚本作用】
读取「意向专用通讯录」备份目录里 **时间戳最新** 的一份 .xlsx 备份，
解析出"学员 ↔ 推荐人"映射，并按学员聚合成 manjike 后端使用的标准 JSON 结构：

    [
      {
        "意向学员微信号": "ceshi1",
        "意向学员微信原始ID": "",
        "是否报名": "未报名",
        "推荐人": [
          {
            "推荐人微信号": "laiyuan1",
            "推荐人微信原始ID": "wxid_laiyuan1",
            "绑定日期": "20260522",
            "绑定状态": "有绑定"
          }
        ]
      }
    ]

【配置驱动】
所有可调项放到同目录的 xlsx_to_prospect_json.config.json，包括：
  - 输入：备份目录、文件名匹配正则、sheet 索引
  - 输出：输出目录、主文件名、是否同时落带时间戳的快照、缩进
  - Excel 表头列名映射（按列名查找，列顺序变化也不影响解析）
  - 输出 JSON 字段名映射
  - 默认值（是否报名 / 解绑日期 / 绑定状态等）

【输出策略（重点）】
默认输出到 **脚本所在目录**（output_dir = "."），不再写到 xlsx 那个网络备份盘。
每次会写两份：
  1) <json_filename>                       例：意向通讯录.json
     永远是最新一次的解析结果，下游脚本可固定读这个文件。
  2) <timestamped_filename_template>       例：意向通讯录_20260418_034510.json
     带时间戳的历史快照，便于回溯（可在配置里关掉）。

【依赖】Python 3.7+ + openpyxl

【模块化】
本脚本采用下划线命名 xlsx_to_prospect_json.py，可被直接 import：
    from xlsx_to_prospect_json import (
        load_config,
        find_latest_xlsx,
        parse_xlsx_to_records,
        run_pipeline,
    )

CLI 用法：
    python xlsx_to_prospect_json.py
    python xlsx_to_prospect_json.py --config some.config.json
    python xlsx_to_prospect_json.py --file <某个具体 xlsx>
    python xlsx_to_prospect_json.py --preview 5    # 只在控制台打印前 5 条，不写文件
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# ============================================================
# 路径常量
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = SCRIPT_DIR / "xlsx_to_prospect_json.config.json"


# ============================================================
# 内置默认配置：当配置文件缺失或某些字段缺省时使用，保证脚本仍可跑
# ============================================================

DEFAULT_CONFIG: Dict[str, Any] = {
    "输入": {
        "backup_dir": r"X:\backup\ProspectiveContacts",
        "filename_pattern": r"^意向专用通讯录导出_(\d{8})_(\d{6})(?:_\d+)?\.xlsx$",
        "sheet_index": 0,
    },
    "输出": {
        "output_dir": ".",
        "json_filename": "意向通讯录.json",
        "save_timestamped_copy": True,
        "timestamped_filename_template": "意向通讯录_{timestamp}.json",
        "indent": 2,
    },
    "Excel表头": {
        "学员微信号列": "意向学员(微信号)",
        "学员微信原始ID列": "意向学员(微信ID)",
        "学员是否删除列": "意向学员(是否删除)",
        "推荐人微信号列": "来源(微信号)",
        "推荐人微信原始ID列": "来源(微信ID)",
        "绑定时间列": "意向学员(添加时间)",
    },
    "输出字段名": {
        "学员对象": {
            "学员微信号": "意向学员微信号",
            "学员微信原始ID": "意向学员微信原始ID",
            "是否报名": "是否报名",
            "推荐人列表": "推荐人",
        },
        "推荐人对象": {
            "推荐人微信号": "推荐人微信号",
            "推荐人微信原始ID": "推荐人微信原始ID",
            "绑定日期": "绑定日期",
            "解绑日期": "解绑日期",
            "绑定状态": "绑定状态",
        },
    },
    "默认值": {
        "是否报名": "未报名",
        "解绑日期": "",
        "绑定状态_未删除": "有绑定",
        "绑定状态_已删除": "",
    },
}


# ============================================================
# 工具函数
# ============================================================

def log(level: str, msg: str) -> None:
    """统一格式打印，加时间戳便于排查。"""
    print(f"{datetime.now().strftime('%H:%M:%S')} [{level}] {msg}", flush=True)


def deep_merge(base: Dict[str, Any], override: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """深合并两层字典：base 是默认值，override 是用户配置。

    - 用户配置中带 _ 开头的 key（注释字段）会被忽略
    - dict 嵌套时递归合并；其他类型直接覆盖
    """
    if not override:
        return dict(base)
    result: Dict[str, Any] = {}
    for key, base_value in base.items():
        if key in override and not (isinstance(key, str) and key.startswith("_")):
            user_value = override[key]
            if isinstance(base_value, dict) and isinstance(user_value, dict):
                result[key] = deep_merge(base_value, user_value)
            else:
                result[key] = user_value
        else:
            result[key] = base_value if not isinstance(base_value, dict) else dict(base_value)
    # 用户配置可能新增 base 里没有的键，也保留下来（保持配置文件的完整原貌）
    for key, user_value in override.items():
        if isinstance(key, str) and key.startswith("_"):
            continue
        if key not in result:
            result[key] = user_value
    return result


def load_config(config_path: Optional[Path] = None) -> Dict[str, Any]:
    """加载配置文件并与默认配置深合并。

    优先级：用户配置 > 内置 DEFAULT_CONFIG。
    配置文件不存在时直接返回默认配置（不报错），方便首次运行。
    """
    if config_path is None:
        config_path = DEFAULT_CONFIG_PATH
    if not config_path.exists():
        log("INFO", f"未找到配置文件 {config_path}，使用内置默认值")
        return dict(DEFAULT_CONFIG)
    try:
        user_cfg = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as ex:
        raise SystemExit(f"[ERROR] 配置文件 JSON 解析失败：{config_path}：{ex}")
    return deep_merge(DEFAULT_CONFIG, user_cfg)


def find_latest_xlsx(directory: Path,
                     filename_pattern: Optional[str] = None) -> Optional[Path]:
    """在目录里找文件名时间戳最新的 .xlsx。

    优先按文件名里的 (YYYYMMDD)_(HHMMSS) 解析，找不到时退化为 mtime 最新者。
    filename_pattern 必须包含两个捕获组（年月日 / 时分秒），否则视为不匹配。
    """
    if not directory.exists() or not directory.is_dir():
        return None
    pattern = re.compile(filename_pattern) if filename_pattern else None

    candidates: List[Tuple[str, Path]] = []
    fallback: List[Path] = []
    for entry in directory.iterdir():
        if not entry.is_file() or entry.suffix.lower() != ".xlsx":
            continue
        ts_key = ""
        if pattern is not None:
            m = pattern.search(entry.name)
            if m and m.lastindex and m.lastindex >= 2:
                # 拼成 YYYYMMDDHHMMSS，便于字符串排序
                ts_key = (m.group(1) or "") + (m.group(2) or "")
        if ts_key:
            candidates.append((ts_key, entry))
        else:
            fallback.append(entry)

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]
    if fallback:
        fallback.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return fallback[0]
    return None


def normalize_value(value: Any) -> str:
    """把 cell 的值统一成 str；None / 空白 → 空字符串。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


# 日期解析格式列表（顺序：最常见的放最前，减少平均 try 次数）
_DATE_FMTS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y/%m/%d",
    "%Y%m%d%H%M%S",
    "%Y%m%d",
)

# 日期解析结果缓存：同一个 xlsx 里绑定日期重复率极高（同月大量相同值），
# 用 dict 缓存避免对重复字符串反复 try 8 种格式。
_date_cache: Dict[str, str] = {}


def parse_bind_date(add_time_str: str) -> str:
    """把"意向学员(添加时间)"转成 YYYYMMDD（带缓存，重复值 O(1)）。"""
    if not add_time_str:
        return ""
    cached = _date_cache.get(add_time_str)
    if cached is not None:
        return cached
    s = add_time_str.strip()
    for fmt in _DATE_FMTS:
        try:
            result = datetime.strptime(s, fmt).strftime("%Y%m%d")
            _date_cache[add_time_str] = result
            return result
        except ValueError:
            continue
    _date_cache[add_time_str] = ""
    return ""


def is_student_active(is_delete_cell: str) -> bool:
    """xlsx 的"是否删除"列：✅=已删除，❌=未删除（参见 db_viewer.py:1738）。"""
    return is_delete_cell != "✅"  # strip 已在 normalize_value 里做过，这里不重复


# ============================================================
# 核心解析（按列名查找，列顺序变化不会影响）
# ============================================================

def _build_column_index(header_row: Sequence[str], column_map: Dict[str, str]) -> Dict[str, int]:
    """根据"业务键 → Excel 列名"映射，反查 Excel 列名在表头里的列索引。

    返回：业务键 → 列索引（0 起）。某列在表头中找不到时，对应业务键不会出现在结果中。
    """
    name_to_idx = {name: idx for idx, name in enumerate(header_row)}
    resolved: Dict[str, int] = {}
    for biz_key, excel_col in column_map.items():
        if excel_col in name_to_idx:
            resolved[biz_key] = name_to_idx[excel_col]
    return resolved


def parse_xlsx_to_records(xlsx_path: Path,
                          config: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """读取 xlsx 并解析成最终的标准 JSON 结构。

    实现策略：
        1) 用 openpyxl 以只读模式打开，避免大文件吃满内存。
        2) 第一行视为表头，按 **列名** 反查列索引（列顺序变化也能正常解析）。
        3) 关键列在表头里缺失时打印 WARN 但不中断；后续行该字段会被空值填充。
        4) 按 (微信ID, 微信号) 聚合学员，多条来源行追加到推荐人数组，保持 xlsx 出现顺序。
    """
    if config is None:
        config = load_config()

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "[ERROR] 缺少依赖 openpyxl，请先安装：\n    pip install openpyxl"
        )

    excel_columns: Dict[str, str] = config["Excel表头"]
    out_student: Dict[str, str] = config["输出字段名"]["学员对象"]
    out_recommender: Dict[str, str] = config["输出字段名"]["推荐人对象"]
    defaults: Dict[str, str] = config["默认值"]

    log("INFO", f"打开 xlsx：{xlsx_path}")
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[int(config["输入"].get("sheet_index", 0))]
        if ws is None:
            raise RuntimeError("xlsx 没有可用的工作表")

        # 读表头并按列名反查索引
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            raise RuntimeError("xlsx 没有表头行")
        header_row = [normalize_value(c) for c in first_row]
        col_idx = _build_column_index(header_row, excel_columns)
        for biz_key, excel_col in excel_columns.items():
            if biz_key not in col_idx:
                log("WARN",
                    f"表头中找不到「{excel_col}」（配置 Excel表头.{biz_key}），"
                    f"该字段将留空")

        # key: (wechat_id, wechat_no) → 学员对象
        students_index: Dict[Tuple[str, str], Dict[str, Any]] = {}
        ordered_students: List[Dict[str, Any]] = []

        skipped_rows = 0
        total_rows = 0

        # ---- 热路径常量提取（避免每行重复 dict 查找）----
        idx_wechat_no   = col_idx.get("学员微信号列")
        idx_wechat_id   = col_idx.get("学员微信原始ID列")
        idx_is_delete   = col_idx.get("学员是否删除列")
        idx_rec_wechat  = col_idx.get("推荐人微信号列")
        idx_rec_oid     = col_idx.get("推荐人微信原始ID列")
        idx_bind_time   = col_idx.get("绑定时间列")

        k_student_wechat   = out_student["学员微信号"]
        k_student_oid      = out_student["学员微信原始ID"]
        k_student_enrolled = out_student["是否报名"]
        k_student_recs     = out_student["推荐人列表"]
        k_rec_wechat       = out_recommender["推荐人微信号"]
        k_rec_oid          = out_recommender["推荐人微信原始ID"]
        k_rec_bind_date    = out_recommender["绑定日期"]
        k_rec_unbind_date  = out_recommender["解绑日期"]
        k_rec_bind_status  = out_recommender["绑定状态"]

        default_enrolled      = defaults.get("是否报名", "未报名")
        default_unbind        = defaults.get("解绑日期", "").strip()
        default_bound_active  = defaults.get("绑定状态_未删除", "").strip()
        default_bound_deleted = defaults.get("绑定状态_已删除", "").strip()

        DELETED_MARK = "✅"  # db_viewer.py 写入时用此符号标记已删除

        n_cols = len(header_row)  # 用于边界检查

        def _cell(raw_row: tuple, idx: Optional[int]) -> str:
            """按预解析列索引取单元格值（内联 normalize 的核心逻辑，避免重建 list）。"""
            if idx is None or idx >= len(raw_row):
                return ""
            v = raw_row[idx]
            if v is None:
                return ""
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d %H:%M:%S")
            return str(v).strip()

        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            # 直接按索引取目标列，不再整行 list comprehension（节省大量无用 normalize）
            student_wechat_no = _cell(raw_row, idx_wechat_no)
            student_wechat_id = _cell(raw_row, idx_wechat_id)

            if not student_wechat_no and not student_wechat_id:
                skipped_rows += 1
                continue

            key = (student_wechat_id, student_wechat_no)
            student = students_index.get(key)
            if student is None:
                student = {
                    k_student_wechat:   student_wechat_no,
                    k_student_oid:      student_wechat_id,
                    k_student_enrolled: default_enrolled,
                    k_student_recs:     [],
                }
                students_index[key] = student
                ordered_students.append(student)

            # 推荐人对象
            recommender: Dict[str, Any] = {
                k_rec_wechat:     _cell(raw_row, idx_rec_wechat),
                k_rec_oid:        _cell(raw_row, idx_rec_oid),
                k_rec_bind_date:  parse_bind_date(_cell(raw_row, idx_bind_time)),
            }
            if default_unbind:
                recommender[k_rec_unbind_date] = default_unbind

            is_deleted_val = _cell(raw_row, idx_is_delete)
            bind_status = default_bound_active if is_deleted_val != DELETED_MARK else default_bound_deleted
            if bind_status:
                recommender[k_rec_bind_status] = bind_status

            student[k_student_recs].append(recommender)

        log("INFO",
            f"解析完成：学员 {len(ordered_students)} 个，处理行 {total_rows} 条，"
            f"跳过空行 {skipped_rows} 条")
        return ordered_students
    finally:
        wb.close()


# ============================================================
# 输出：默认到脚本所在目录 + 时间戳快照
# ============================================================

def _resolve_output_dir(output_dir_value: str) -> Path:
    """把配置里的 output_dir 解析成绝对路径。

    - 绝对路径 → 直接使用
    - "." 或相对路径 → 相对脚本所在目录
    """
    p = Path(output_dir_value)
    if p.is_absolute():
        return p
    return (SCRIPT_DIR / p).resolve()


def write_outputs(records: List[Dict[str, Any]],
                  config: Dict[str, Any],
                  source_xlsx: Optional[Path] = None) -> List[Path]:
    """把解析结果写到磁盘，返回所有生成文件的路径列表。

    永远写主文件 <output_dir>/<json_filename>；
    若 save_timestamped_copy=True，再写一份带时间戳的快照便于回溯。
    """
    output_cfg = config["输出"]
    out_dir = _resolve_output_dir(output_cfg.get("output_dir", "."))
    out_dir.mkdir(parents=True, exist_ok=True)

    indent = output_cfg.get("indent", 2)
    written: List[Path] = []

    # 1) 主文件：用流式写入（先 json.dumps → 直接 write bytes），
    #    比 write_text(json.dumps(...)) 少一次全量字符串在内存里驻留。
    main_path = out_dir / output_cfg.get("json_filename", "意向通讯录.json")
    _write_json(main_path, records, indent)
    written.append(main_path)
    log("INFO", f"已写入主文件：{main_path}（{main_path.stat().st_size} B）")

    # 2) 时间戳快照（可选）
    if output_cfg.get("save_timestamped_copy", True):
        timestamp = _extract_timestamp(source_xlsx) or datetime.now().strftime("%Y%m%d_%H%M%S")
        template = output_cfg.get("timestamped_filename_template",
                                  "意向通讯录_{timestamp}.json")
        snapshot_name = template.format(timestamp=timestamp)
        snapshot_path = out_dir / snapshot_name
        snapshot_path.parent.mkdir(parents=True, exist_ok=True)
        # 快照直接硬链接到主文件（同磁盘），避免第二次全量序列化
        try:
            import os as _os
            if snapshot_path.exists():
                snapshot_path.unlink()
            _os.link(main_path, snapshot_path)
        except (OSError, NotImplementedError):
            # 跨设备/跨文件系统时硬链接失败，退化为复制
            import shutil as _shutil
            _shutil.copy2(main_path, snapshot_path)
        written.append(snapshot_path)
        log("INFO", f"已写入时间戳快照：{snapshot_path}")

    return written


def _write_json(path: Path, data: Any, indent: int) -> None:
    """把 data 序列化为 JSON 并写入文件。

    用 ensure_ascii=False 保留中文字符；
    直接以 bytes 模式写出（encode 一次），避免 str → file 再转 bytes 的双重开销。
    """
    path.write_bytes(json.dumps(data, ensure_ascii=False, indent=indent).encode("utf-8"))


_TIMESTAMP_FROM_NAME = re.compile(r"(\d{8})_(\d{6})")


def _extract_timestamp(source: Optional[Path]) -> str:
    """从源 xlsx 文件名里抽取 YYYYMMDD_HHMMSS；抽不到则返回空串。"""
    if source is None:
        return ""
    m = _TIMESTAMP_FROM_NAME.search(source.name)
    if not m:
        return ""
    return f"{m.group(1)}_{m.group(2)}"


# ============================================================
# 公开 API：run_pipeline()
# ============================================================

def run_pipeline(config_path: Optional[Path] = None,
                 explicit_file: Optional[Path] = None,
                 preview: int = 0) -> int:
    """端到端：读配置 → 找最新 xlsx → 解析 → 写 JSON（含时间戳快照）。

    参数：
        config_path:   配置文件路径；不传则用同目录默认配置
        explicit_file: 直接指定具体 xlsx 文件，跳过"找最新"
        preview:       >0 时仅打印前 N 条到控制台，不写文件

    返回：0 = 成功；非 0 = 失败码
    """
    config = load_config(config_path)

    # === 1) 选定要解析的 xlsx ===
    if explicit_file is not None:
        xlsx_path = explicit_file
        if not xlsx_path.exists():
            log("ERROR", f"指定的 xlsx 不存在：{xlsx_path}")
            return 2
    else:
        backup_dir = Path(config["输入"]["backup_dir"])
        xlsx_path = find_latest_xlsx(
            backup_dir,
            filename_pattern=config["输入"].get("filename_pattern"),
        )
        if xlsx_path is None:
            log("ERROR",
                f"目录里没有可用 .xlsx：{backup_dir}（请检查目录是否存在 / 是否有备份文件）")
            return 2
        log("INFO", f"备份目录：{backup_dir}")
        log("INFO", f"挑选最新 xlsx：{xlsx_path.name}")

    # === 2) 解析 ===
    records = parse_xlsx_to_records(xlsx_path, config=config)
    if not records:
        log("WARN", "解析结果为空（xlsx 没有数据行）")
        return 0

    # === 3) 预览 / 写盘 ===
    if preview > 0:
        log("INFO", f"预览前 {preview} 条（不写文件）：")
        print(json.dumps(records[:preview], ensure_ascii=False, indent=2))
        return 0

    write_outputs(records, config, source_xlsx=xlsx_path)
    return 0


# ============================================================
# CLI 入口
# ============================================================

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="意向通讯录 xlsx → 标准 JSON（配置驱动）",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--config", type=str, default=None,
                        help="配置文件路径，默认为脚本同目录的 "
                             "xlsx_to_prospect_json.config.json")
    parser.add_argument("--file", type=str, default=None,
                        help="直接指定具体 .xlsx 文件，跳过\"找最新\"")
    parser.add_argument("--preview", type=int, default=0,
                        help=">0 时只打印前 N 条到控制台，不写文件，便于快速校验")
    return parser.parse_args()


def _cli_main() -> int:
    args = _parse_args()
    # Windows 控制台 GBK → UTF-8
    if os.name == "nt":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        except Exception:
            pass

    return run_pipeline(
        config_path=Path(args.config) if args.config else None,
        explicit_file=Path(args.file) if args.file else None,
        preview=args.preview,
    )


if __name__ == "__main__":
    exit_code = 0
    try:
        exit_code = _cli_main()
    except KeyboardInterrupt:
        log("CANCELLED", "用户中断（Ctrl+C）")
        exit_code = 130
    except SystemExit as ex:
        exit_code = ex.code if isinstance(ex.code, int) else 1
    except Exception as ex:
        import traceback
        traceback.print_exc()
        log("ERROR", f"未处理异常：{ex}")
        exit_code = 1
    sys.exit(exit_code)
