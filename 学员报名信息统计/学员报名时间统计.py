import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
import matplotlib.pyplot as plt  # pyright: ignore[reportMissingImports]
import numpy as np
from matplotlib.widgets import Slider  # pyright: ignore[reportMissingImports]

def parse_time(cell_value):
    if pd.isna(cell_value):
        return None, False
    
    cell_str = str(cell_value).strip()
    
    if cell_str == '❌':
        return '❌', True
    
    patterns = [
        (r'(\d{4})[/.-](\d{1,2})[/.-](\d{1,2})', lambda m: f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"),
        (r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', lambda m: f"{m.group(3)}/{int(m.group(1))}/{int(m.group(2))}"),
        (r'(\d{4})年(\d{1,2})月(\d{1,2})日', lambda m: f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"),
        (r'(\d{1,2})月(\d{1,2})日(\d{4})年', lambda m: f"{m.group(3)}/{int(m.group(1))}/{int(m.group(2))}"),
    ]
    
    for pattern, formatter in patterns:
        match = re.search(pattern, cell_str)
        if match:
            return formatter(match), True
    
    try:
        dt = pd.to_datetime(cell_value)
        return dt.strftime('%Y/%m/%d'), True
    except:
        pass
    
    return None, False

def main():
    file_path = r'C:\Users\LENOVO\Desktop\报名录入.xlsx'
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb['报名录入']
    
    headers = [cell.value for cell in ws[2]]
    
    new_old_col_idx = None
    time_col_idx = None
    
    for idx, header in enumerate(headers):
        if header and '新旧编号' in str(header):
            new_old_col_idx = idx
        if header and '时间' in str(header):
            time_col_idx = idx
    
    if new_old_col_idx is None or time_col_idx is None:
        print("未找到新旧编号列或时间列")
        return
    
    daily_new, daily_renewal, daily_total = {}, {}, {}
    weekly_new, weekly_renewal, weekly_total = {}, {}, {}
    monthly_new, monthly_renewal, monthly_total = {}, {}, {}
    
    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if row[new_old_col_idx] is None or row[time_col_idx] is None:
            continue
        
        new_old = str(row[new_old_col_idx]).strip() if row[new_old_col_idx] else ''
        time_val, valid = parse_time(row[time_col_idx])
        
        if not valid or time_val == '❌':
            continue
        
        try:
            dt = datetime.strptime(time_val, '%Y/%m/%d')
        except:
            continue
        
        date_str = time_val
        week_str = f"{dt.year}-W{dt.isocalendar()[1]:02d}"
        month_str = dt.strftime('%Y/%m')
        
        if new_old == '新':
            daily_new[date_str] = daily_new.get(date_str, 0) + 1
            weekly_new[week_str] = weekly_new.get(week_str, 0) + 1
            monthly_new[month_str] = monthly_new.get(month_str, 0) + 1
        elif new_old == '旧':
            daily_renewal[date_str] = daily_renewal.get(date_str, 0) + 1
            weekly_renewal[week_str] = weekly_renewal.get(week_str, 0) + 1
            monthly_renewal[month_str] = monthly_renewal.get(month_str, 0) + 1
    
    for date_str in set(daily_new.keys()) | set(daily_renewal.keys()):
        daily_total[date_str] = daily_new.get(date_str, 0) + daily_renewal.get(date_str, 0)
    
    for week_str in set(weekly_new.keys()) | set(weekly_renewal.keys()):
        weekly_total[week_str] = weekly_new.get(week_str, 0) + weekly_renewal.get(week_str, 0)
    
    for month_str in set(monthly_new.keys()) | set(monthly_renewal.keys()):
        monthly_total[month_str] = monthly_new.get(month_str, 0) + monthly_renewal.get(month_str, 0)
    
    print("=" * 80)
    print("每日统计")
    print("=" * 80)
    print(f"{'日期':<12} {'新增':<10} {'续费':<10} {'总计':<10}")
    print("-" * 42)
    for date_str in sorted(daily_new.keys() | daily_renewal.keys()):
        print(f"{date_str:<12} {daily_new.get(date_str, 0):<10} {daily_renewal.get(date_str, 0):<10} {daily_total.get(date_str, 0):<10}")
    
    print("\n" + "=" * 80)
    print("每周统计")
    print("=" * 80)
    print(f"{'周次':<15} {'新增':<10} {'续费':<10} {'总计':<10}")
    print("-" * 45)
    for week_str in sorted(weekly_new.keys() | weekly_renewal.keys()):
        print(f"{week_str:<15} {weekly_new.get(week_str, 0):<10} {weekly_renewal.get(week_str, 0):<10} {weekly_total.get(week_str, 0):<10}")
    
    print("\n" + "=" * 80)
    print("每月统计")
    print("=" * 80)
    print(f"{'月份':<12} {'新增':<10} {'续费':<10} {'总计':<10}")
    print("-" * 42)
    for month_str in sorted(monthly_new.keys() | monthly_renewal.keys()):
        print(f"{month_str:<12} {monthly_new.get(month_str, 0):<10} {monthly_renewal.get(month_str, 0):<10} {monthly_total.get(month_str, 0):<10}")
    
    print("\n" + "=" * 80)
    print("总计")
    print("=" * 80)
    print(f"新增人数: {sum(daily_new.values())}")
    print(f"续费人数: {sum(daily_renewal.values())}")
    print(f"总计人数: {sum(daily_total.values())}")
    
    plot_charts(daily_new, daily_renewal, daily_total, 
                weekly_new, weekly_renewal, weekly_total,
                monthly_new, monthly_renewal, monthly_total)

def plot_charts(daily_new, daily_renewal, daily_total,
                weekly_new, weekly_renewal, weekly_total,
                monthly_new, monthly_renewal, monthly_total):
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    
    fig, axes = plt.subplots(3, 1, figsize=(14, 10))
    fig.suptitle('学员报名时间趋势分析', fontsize=16, fontweight='bold', y=0.995)
    
    dates = sorted(daily_new.keys() | daily_renewal.keys())
    new_vals = [daily_new.get(d, 0) for d in dates]
    renewal_vals = [daily_renewal.get(d, 0) for d in dates]
    total_vals = [daily_total.get(d, 0) for d in dates]
    
    line1, = axes[0].plot(dates, new_vals, marker='o', label='新增', linewidth=2, markersize=4, color='#2E86B1')
    line2, = axes[0].plot(dates, renewal_vals, marker='s', label='续费', linewidth=2, markersize=4, color='#E67E22')
    line3, = axes[0].plot(dates, total_vals, marker='^', label='总计', linewidth=2, markersize=4, color='#27AE60')
    axes[0].set_title('每日报名趋势', fontsize=14, fontweight='bold')
    axes[0].set_xlabel('日期')
    axes[0].set_ylabel('人数')
    axes[0].legend(loc='upper left')
    axes[0].grid(True, alpha=0.3)
    axes[0].set_xlim(dates[0], dates[-1])
    
    weeks = sorted(weekly_new.keys() | weekly_renewal.keys())
    new_week_vals = [weekly_new.get(w, 0) for w in weeks]
    renewal_week_vals = [weekly_renewal.get(w, 0) for w in weeks]
    total_week_vals = [weekly_total.get(w, 0) for w in weeks]
    
    line4, = axes[1].plot(weeks, new_week_vals, marker='o', label='新增', linewidth=2, markersize=4, color='#2E86B1')
    line5, = axes[1].plot(weeks, renewal_week_vals, marker='s', label='续费', linewidth=2, markersize=4, color='#E67E22')
    line6, = axes[1].plot(weeks, total_week_vals, marker='^', label='总计', linewidth=2, markersize=4, color='#27AE60')
    axes[1].set_title('每周报名趋势', fontsize=14, fontweight='bold')
    axes[1].set_xlabel('周次')
    axes[1].set_ylabel('人数')
    axes[1].legend(loc='upper left')
    axes[1].grid(True, alpha=0.3)
    axes[1].set_xlim(weeks[0], weeks[-1])
    
    months = sorted(monthly_new.keys() | monthly_renewal.keys())
    new_month_vals = [monthly_new.get(m, 0) for m in months]
    renewal_month_vals = [monthly_renewal.get(m, 0) for m in months]
    total_month_vals = [monthly_total.get(m, 0) for m in months]
    
    line7, = axes[2].plot(months, new_month_vals, marker='o', label='新增', linewidth=2, markersize=4, color='#2E86B1')
    line8, = axes[2].plot(months, renewal_month_vals, marker='s', label='续费', linewidth=2, markersize=4, color='#E67E22')
    line9, = axes[2].plot(months, total_month_vals, marker='^', label='总计', linewidth=2, markersize=4, color='#27AE60')
    axes[2].set_title('每月报名趋势', fontsize=14, fontweight='bold')
    axes[2].set_xlabel('月份')
    axes[2].set_ylabel('人数')
    axes[2].legend(loc='upper left')
    axes[2].grid(True, alpha=0.3)
    axes[2].set_xlim(months[0], months[-1])
    
    plt.tight_layout()
    
    plt.subplots_adjust(left=0.08, right=0.98, top=0.96, bottom=0.08, hspace=0.3)
    
    slider_ax = plt.axes([0.1, 0.02, 0.8, 0.03], facecolor='#f0f0f0')
    slider = Slider(slider_ax, '时间范围', 0, 100, valinit=100, valstep=1)
    
    original_xlims = [
        (axes[0].get_xlim(), axes[0].get_ylim()),
        (axes[1].get_xlim(), axes[1].get_ylim()),
        (axes[2].get_xlim(), axes[2].get_ylim())
    ]
    
    drag_state = {
        'active': False,
        'axis': None,
        'x0': None,
        'xlim0': None,
        'last_x': None
    }
    
    def on_press(event):
        if event.inaxes is None or event.button != 1:
            return
        
        for i, ax in enumerate(axes):
            if event.inaxes == ax:
                drag_state['active'] = True
                drag_state['axis'] = ax
                drag_state['x0'] = event.xdata
                drag_state['xlim0'] = ax.get_xlim()
                drag_state['last_x'] = event.xdata
                ax.figure.canvas.cursor().setCursor(8)
                break
    
    def on_move(event):
        if not drag_state['active'] or event.inaxes is None:
            return
        
        if drag_state['axis'] and event.inaxes == drag_state['axis']:
            current_x = event.xdata
            if drag_state['last_x'] is not None and current_x is not None:
                xlim = drag_state['xlim0']
                dx = current_x - drag_state['last_x']
                new_xlim = (xlim[0] - dx, xlim[1] - dx)
                
                range_width = xlim[1] - xlim[0]
                if drag_state['axis'] == axes[0]:
                    data_range = (dates[0], dates[-1])
                elif drag_state['axis'] == axes[1]:
                    data_range = (weeks[0], weeks[-1])
                else:
                    data_range = (months[0], months[-1])
                
                min_span = (data_range[1] - data_range[0]) * 0.1
                if range_width < min_span:
                    range_width = min_span
                
                if new_xlim[0] >= data_range[0] - range_width * 0.5 and new_xlim[1] <= data_range[1] + range_width * 0.5:
                    drag_state['axis'].set_xlim(new_xlim)
                    drag_state['last_x'] = current_x
                    drag_state['axis'].figure.canvas.draw_idle()
    
    def on_release(event):
        if drag_state['active'] and drag_state['axis']:
            drag_state['axis'].figure.canvas.cursor().setCursor(0)
            drag_state['active'] = False
            drag_state['axis'] = None
            drag_state['last_x'] = None
    
    def on_scroll(event):
        if event.inaxes is None:
            return
        
        for ax in axes:
            if event.inaxes == ax:
                xlim = ax.get_xlim()
                range_width = xlim[1] - xlim[0]
                scroll_factor = 0.1
                
                if event.button == 'up':
                    new_range = range_width * (1 - scroll_factor)
                    center = (xlim[0] + xlim[1]) / 2
                else:
                    new_range = range_width * (1 + scroll_factor)
                    center = (xlim[0] + xlim[1]) / 2
                
                if new_range < range_width * 0.05:
                    new_range = range_width * 0.05
                
                if ax == axes[0]:
                    data_range = (dates[0], dates[-1])
                elif ax == axes[1]:
                    data_range = (weeks[0], weeks[-1])
                else:
                    data_range = (months[0], months[-1])
                
                min_span = (data_range[1] - data_range[0]) * 0.1
                if new_range < min_span:
                    new_range = min_span
                
                half_range = new_range / 2
                new_xlim = (center - half_range, center + half_range)
                
                if new_xlim[0] < data_range[0]:
                    new_xlim = (data_range[0], data_range[0] + new_range)
                if new_xlim[1] > data_range[1]:
                    new_xlim = (data_range[1] - new_range, data_range[1])
                
                ax.set_xlim(new_xlim)
                ax.figure.canvas.draw_idle()
                break
    
    fig.canvas.mpl_connect('button_press_event', on_press)
    fig.canvas.mpl_connect('motion_notify_event', on_move)
    fig.canvas.mpl_connect('button_release_event', on_release)
    fig.canvas.mpl_connect('scroll_event', on_scroll)
    
    def update_slider(val):
        percentage = slider.val / 100.0
        
        for i, ax in enumerate(axes):
            if i == 0:
                data_range = (dates[0], dates[-1])
            elif i == 1:
                data_range = (weeks[0], weeks[-1])
            else:
                data_range = (months[0], months[-1])
            
            range_width = data_range[1] - data_range[0]
            new_width = range_width * (0.1 + percentage * 0.9)
            center = (data_range[0] + data_range[1]) / 2
            new_xlim = (center - new_width / 2, center + new_width / 2)
            
            if new_xlim[0] < data_range[0]:
                new_xlim = (data_range[0], data_range[0] + new_width)
            if new_xlim[1] > data_range[1]:
                new_xlim = (data_range[1] - new_width, data_range[1])
            
            ax.set_xlim(new_xlim)
        
        fig.canvas.draw_idle()
    
    slider.on_changed(update_slider)
    
    plt.savefig('d:\\桌面文件\\新建文件夹\\uploads\\报名趋势图.png', dpi=300, bbox_inches='tight')
    print("\n图表已保存至: d:\\桌面文件\\新建文件夹\\uploads\\报名趋势图.png")
    print("\n交互功能说明:")
    print("- 鼠标左键拖动: 水平平移图表")
    print("- 滚轮滚动: 放大/缩小图表")
    print("- 底部滑块: 调整显示范围")
    plt.show()

if __name__ == '__main__':
    main()
