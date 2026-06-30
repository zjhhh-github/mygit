import openpyxl
import re

file_path = r"C:\Users\LENOVO\Desktop\CT外教充值记录.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

rows = []
date_pattern = re.compile(r'^\d{8}$')

max_row = sheet.max_row
max_col = sheet.max_column

for row_idx in range(1, max_row + 1):
    row_data = []
    for col_idx in range(1, max_col + 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        row_data.append(cell_value)
    rows.append({
        'row_idx': row_idx,
        'values': row_data
    })

date_rows = []
for i, row_data in enumerate(rows):
    values = row_data['values']
    if len(values) >= 4:
        a_val = str(values[0]).strip() if values[0] else ''
        b_val = str(values[1]).strip() if values[1] else ''
        c_val = str(values[2]).strip() if values[2] else ''
        d_val = str(values[3]).strip() if values[3] else ''
        
        if (not b_val or b_val == 'None') and (not c_val or c_val == 'None') and (not d_val or d_val == 'None'):
            if date_pattern.match(a_val):
                formatted_date = f"{a_val[:4]}-{a_val[4:6]}-{a_val[6:]}"
                date_rows.append({
                    'index': i,
                    'date': formatted_date,
                    'row_idx': row_data['row_idx']
                })

for i in range(len(date_rows) - 1):
    start_idx = date_rows[i]['index'] + 1
    end_idx = date_rows[i + 1]['index']
    current_date = date_rows[i]['date']
    
    for j in range(start_idx, end_idx):
        if j < len(rows):
            values = rows[j]['values']
            while len(values) < 4:
                values.append(None)
            values[2] = current_date
            rows[j]['values'] = values

if len(date_rows) > 0:
    last_date = date_rows[-1]['date']
    start_idx = date_rows[-1]['index'] + 1
    for j in range(start_idx, len(rows)):
        values = rows[j]['values']
        while len(values) < 4:
            values.append(None)
        values[2] = last_date
        rows[j]['values'] = values

delete_row_indices = set([row['row_idx'] for row in date_rows])

new_rows = []
for row_data in rows:
    if row_data['row_idx'] not in delete_row_indices:
        new_rows.append(row_data)

for new_idx, row_data in enumerate(new_rows, start=1):
    for col_idx, value in enumerate(row_data['values'], start=1):
        sheet.cell(row=new_idx, column=col_idx, value=value)

while sheet.max_row > len(new_rows):
    sheet.delete_rows(sheet.max_row)

output_path = r"C:\Users\LENOVO\Desktop\CT外教充值记录_处理后.xlsx"
workbook.save(output_path)

print(f"处理完成！")
print(f"共找到 {len(date_rows)} 个日期标记")
print(f"原始行数: {len(rows)}, 处理后行数: {len(new_rows)}")
print(f"已删除日期行")
print(f"处理后的文件保存至: {output_path}")
